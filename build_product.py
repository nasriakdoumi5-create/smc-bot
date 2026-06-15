"""
Restaurant & Cafe Management Template Builder
Builds a full professional Excel file ready for Google Sheets
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

# ══ Colors ══════════════════════════════════════════
C_DARK    = "1A1A2E"   # dark navy
C_PRIMARY = "16213E"   # primary
C_ACCENT  = "E94560"   # red accent
C_GREEN   = "0F9B58"   # green
C_YELLOW  = "FFD700"   # input cells
C_LIGHT   = "F5F5F5"   # light bg
C_WHITE   = "FFFFFF"
C_ORANGE  = "FF6B35"
C_BLUE    = "4A90D9"
C_HEADER  = "0D3B66"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="888888")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(ws, row, col, value, bg=C_HEADER, fg=C_WHITE, size=12, bold=True, width=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = font(bold=bold, size=size, color=fg)
    cell.alignment = align()
    cell.border = border_thin()
    return cell

def style_cell(ws, row, col, value=None, bg=C_WHITE, fg="000000",
               bold=False, h="center", wrap=False, border=True, size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = font(bold=bold, size=size, color=fg)
    cell.alignment = align(h=h, wrap=wrap)
    if border:
        cell.border = border_thin()
    return cell

def input_cell(ws, row, col, value=None, hint=""):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(C_YELLOW)
    cell.font = font(size=11)
    cell.alignment = align(h="center")
    cell.border = border_medium()
    return cell

# ══════════════════════════════════════════════════════
#  TAB 1 — DASHBOARD
# ══════════════════════════════════════════════════════
def build_dashboard(wb):
    ws = wb.active
    ws.title = "📊 Dashboard"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_ACCENT

    # column widths
    widths = [2, 20, 18, 18, 18, 18, 18, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 50
    ws.row_dimensions[3].height = 15

    # ── Title Banner ──────────────────────────────────
    ws.merge_cells("B2:G2")
    cell = ws["B2"]
    cell.value = "🍽️  RESTAURANT & CAFE MANAGER"
    cell.fill = fill(C_DARK)
    cell.font = font(bold=True, size=22, color=C_YELLOW)
    cell.alignment = align()

    ws.merge_cells("B3:G3")
    sub = ws["B3"]
    sub.value = "Smart Business Automation Template"
    sub.fill = fill(C_DARK)
    sub.font = font(size=11, color="AAAAAA", italic=True)
    sub.alignment = align()

    # ── KPI Cards ─────────────────────────────────────
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 35
    ws.row_dimensions[7].height = 40
    ws.row_dimensions[8].height = 25

    kpis = [
        ("💰 TODAY'S SALES",   "='📈 Daily Sales'!C32",  C_GREEN,  "B"),
        ("📦 LOW STOCK ITEMS", "='📦 Inventory'!H2",     C_ACCENT, "D"),
        ("👥 STAFF TODAY",     "='👥 Staff Schedule'!H2", C_BLUE,  "F"),
    ]

    for title, formula, color, col in kpis:
        ws.merge_cells(f"{col}5:{chr(ord(col)+1)}5")
        ws.merge_cells(f"{col}6:{chr(ord(col)+1)}6")
        ws.merge_cells(f"{col}7:{chr(ord(col)+1)}7")
        ws.merge_cells(f"{col}8:{chr(ord(col)+1)}8")

        t = ws[f"{col}5"]
        t.fill = fill(color)
        t.font = font(bold=True, size=9, color=C_WHITE)
        t.alignment = align()
        t.value = title

        v = ws[f"{col}7"]
        v.fill = fill(color)
        v.font = font(bold=True, size=26, color=C_WHITE)
        v.alignment = align()
        v.value = formula

        for r in [6, 8]:
            c = ws[f"{col}{r}"]
            c.fill = fill(color)

    # ── Weekly Summary Table ───────────────────────────
    ws.row_dimensions[10].height = 25
    ws.merge_cells("B10:G10")
    sec = ws["B10"]
    sec.value = "📅  WEEKLY SALES SUMMARY"
    sec.fill = fill(C_PRIMARY)
    sec.font = font(bold=True, size=13, color=C_WHITE)
    sec.alignment = align()

    headers = ["Day", "Sales ($)", "Expenses ($)", "Profit ($)", "Orders", "Avg Order ($)"]
    days    = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    for i, h in enumerate(headers, 2):
        style_header(ws, 11, i, h, bg=C_HEADER)

    for r, day in enumerate(days, 12):
        ws.row_dimensions[r].height = 22
        style_cell(ws, r, 2, day, bold=True, bg=C_LIGHT if r % 2 == 0 else C_WHITE)
        for c in range(3, 8):
            if c == 5:  # Profit
                cell = ws.cell(row=r, column=c,
                               value=f"=C{r}-D{r}")
                cell.fill = fill(C_LIGHT if r % 2 == 0 else C_WHITE)
                cell.font = font(color=C_GREEN, bold=True)
                cell.alignment = align()
                cell.border = border_thin()
                cell.number_format = '"$"#,##0.00'
            elif c == 7:  # Avg Order
                cell = ws.cell(row=r, column=c,
                               value=f"=IF(F{r}=0,0,C{r}/F{r})")
                cell.fill = fill(C_LIGHT if r % 2 == 0 else C_WHITE)
                cell.font = font()
                cell.alignment = align()
                cell.border = border_thin()
                cell.number_format = '"$"#,##0.00'
            else:
                ic = input_cell(ws, r, c, 0)
                ic.number_format = '"$"#,##0.00' if c in [3, 4] else "0"

    # Totals row
    ws.row_dimensions[19].height = 28
    style_cell(ws, 19, 2, "TOTAL", bold=True, bg=C_DARK, fg=C_WHITE)
    for c in range(3, 8):
        col_l = get_column_letter(c)
        cell = ws.cell(row=19, column=c,
                       value=f"=SUM({col_l}12:{col_l}18)")
        cell.fill = fill(C_DARK)
        cell.font = font(bold=True, color=C_YELLOW)
        cell.alignment = align()
        cell.border = border_thin()
        if c != 6:
            cell.number_format = '"$"#,##0.00'

    # ── Tips ──────────────────────────────────────────
    ws.row_dimensions[21].height = 20
    ws.merge_cells("B21:G21")
    tip = ws["B21"]
    tip.value = "💡  TIP: Fill yellow cells with your daily data. Green cells calculate automatically."
    tip.fill = fill("FFF9C4")
    tip.font = font(size=10, color="7B6000", italic=True)
    tip.alignment = align(h="left")


# ══════════════════════════════════════════════════════
#  TAB 2 — INVENTORY
# ══════════════════════════════════════════════════════
def build_inventory(wb):
    ws = wb.create_sheet("📦 Inventory")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_GREEN

    widths = [2, 25, 15, 12, 12, 12, 15, 15, 15, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.row_dimensions[1].height = 40
    ws.merge_cells("B1:I1")
    t = ws["B1"]
    t.value = "📦  INVENTORY TRACKER"
    t.fill = fill(C_GREEN)
    t.font = font(bold=True, size=18, color=C_WHITE)
    t.alignment = align()

    # Low stock counter formula in H2 (referenced by dashboard)
    ws["H2"] = '=COUNTIF(F3:F52,"<="&G3:G52)'
    ws["H2"].font = font(size=1, color=C_WHITE)  # hidden helper cell

    # Headers
    headers = ["Item Name", "Category", "Unit", "Current Stock",
               "Min Stock", "Status", "Unit Cost ($)", "Total Value ($)"]
    for i, h in enumerate(headers, 2):
        style_header(ws, 3, i, h)

    # Sample data
    items = [
        ("Coffee Beans",    "Beverages", "kg",  50,  10),
        ("Milk",            "Beverages", "L",   30,  10),
        ("Sugar",           "Dry Goods", "kg",  20,   5),
        ("Flour",           "Dry Goods", "kg",  15,   5),
        ("Chicken",         "Proteins",  "kg",  10,   3),
        ("Tomatoes",        "Vegetables","kg",   8,   3),
        ("Olive Oil",       "Condiments","L",    5,   2),
        ("Cheese",          "Dairy",     "kg",   6,   2),
        ("Bread Loaves",    "Bakery",    "pcs", 20,   5),
        ("Disposable Cups", "Supplies",  "pcs",200,  50),
    ]

    for r, (name, cat, unit, stock, min_s) in enumerate(items, 4):
        ws.row_dimensions[r].height = 22
        bg = C_LIGHT if r % 2 == 0 else C_WHITE

        style_cell(ws, r, 2, name, bold=True, bg=bg, h="left")
        style_cell(ws, r, 3, cat, bg=bg)
        style_cell(ws, r, 4, unit, bg=bg)
        input_cell(ws, r, 5, stock)
        input_cell(ws, r, 6, min_s)

        # Status formula
        status_cell = ws.cell(row=r, column=7,
            value=f'=IF(E{r}<=F{r},"⚠️ LOW","✅ OK")')
        status_cell.fill = fill(bg)
        status_cell.font = font(bold=True)
        status_cell.alignment = align()
        status_cell.border = border_thin()

        input_cell(ws, r, 8, 0).number_format = '"$"#,##0.00'

        # Total value
        tv = ws.cell(row=r, column=9, value=f"=E{r}*H{r}")
        tv.fill = fill(bg)
        tv.font = font(bold=True, color=C_GREEN)
        tv.alignment = align()
        tv.border = border_thin()
        tv.number_format = '"$"#,##0.00'

    # Add blank rows for more items
    for r in range(14, 53):
        ws.row_dimensions[r].height = 22
        bg = C_LIGHT if r % 2 == 0 else C_WHITE
        for c in range(2, 10):
            style_cell(ws, r, c, bg=bg)
        for c in [5, 6, 8]:
            input_cell(ws, r, c)
        ws.cell(row=r, column=7, value=f'=IF(E{r}="","",IF(E{r}<=F{r},"⚠️ LOW","✅ OK"))')
        ws.cell(row=r, column=9, value=f"=IF(E{r}=\"\",\"\",E{r}*H{r})")

    # Summary
    ws.row_dimensions[54].height = 28
    ws.merge_cells("B54:H54")
    style_cell(ws, 54, 2, "TOTAL INVENTORY VALUE", bold=True, bg=C_DARK, fg=C_WHITE)
    tv_total = ws.cell(row=54, column=9, value="=SUM(I4:I53)")
    tv_total.fill = fill(C_DARK)
    tv_total.font = font(bold=True, color=C_YELLOW, size=14)
    tv_total.alignment = align()
    tv_total.number_format = '"$"#,##0.00'


# ══════════════════════════════════════════════════════
#  TAB 3 — DAILY SALES
# ══════════════════════════════════════════════════════
def build_daily_sales(wb):
    ws = wb.create_sheet("📈 Daily Sales")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_BLUE

    widths = [2, 8, 25, 12, 12, 12, 15, 12, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 40
    ws.merge_cells("B1:H1")
    t = ws["B1"]
    t.value = "📈  DAILY SALES TRACKER"
    t.fill = fill(C_BLUE)
    t.font = font(bold=True, size=18, color=C_WHITE)
    t.alignment = align()

    # C32 referenced by dashboard for today's total
    ws["C32"] = "=SUM(E4:E31)"
    ws["C32"].font = font(size=1, color=C_WHITE)

    headers = ["#", "Item / Service", "Category", "Qty", "Price ($)", "Total ($)", "Payment", "Time"]
    for i, h in enumerate(headers, 2):
        style_header(ws, 3, i, h)

    categories = ["Food", "Beverage", "Dessert", "Takeaway", "Delivery"]
    payments   = ["Cash", "Card", "Online", "Other"]

    dv_cat = DataValidation(type="list",
                             formula1='"Food,Beverage,Dessert,Takeaway,Delivery"',
                             allow_blank=True)
    dv_pay = DataValidation(type="list",
                             formula1='"Cash,Card,Online,Other"',
                             allow_blank=True)
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_pay)

    for r in range(4, 32):
        ws.row_dimensions[r].height = 22
        bg = C_LIGHT if r % 2 == 0 else C_WHITE

        style_cell(ws, r, 2, r - 3, bg=bg)

        for c in [3, 4]:
            ic = input_cell(ws, r, c)
            if c == 4:
                dv_cat.add(ic)

        for c in [5, 6]:
            ic = input_cell(ws, r, c, 0)
            ic.number_format = '"$"#,##0.00'

        total = ws.cell(row=r, column=7,
                        value=f"=IF(E{r}=0,\"\",E{r}*F{r})")
        total.fill = fill(bg)
        total.font = font(bold=True, color=C_GREEN)
        total.alignment = align()
        total.border = border_thin()
        total.number_format = '"$"#,##0.00'

        pm = input_cell(ws, r, 8)
        dv_pay.add(pm)

        input_cell(ws, r, 9)

    # Totals
    ws.row_dimensions[33].height = 28
    style_cell(ws, 33, 2, "", bg=C_DARK)
    ws.merge_cells("C33:F33")
    style_cell(ws, 33, 3, "DAILY TOTAL", bold=True, bg=C_DARK, fg=C_WHITE)
    total_cell = ws.cell(row=33, column=7, value="=SUM(G4:G31)")
    total_cell.fill = fill(C_DARK)
    total_cell.font = font(bold=True, color=C_YELLOW, size=14)
    total_cell.alignment = align()
    total_cell.number_format = '"$"#,##0.00'


# ══════════════════════════════════════════════════════
#  TAB 4 — STAFF SCHEDULE
# ══════════════════════════════════════════════════════
def build_staff(wb):
    ws = wb.create_sheet("👥 Staff Schedule")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_ORANGE

    widths = [2, 20, 12, 10, 10, 10, 10, 10, 10, 10, 15, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 40
    ws.merge_cells("B1:K1")
    t = ws["B1"]
    t.value = "👥  STAFF WEEKLY SCHEDULE"
    t.fill = fill(C_ORANGE)
    t.font = font(bold=True, size=18, color=C_WHITE)
    t.alignment = align()

    # H2 referenced by dashboard
    ws["H2"] = '=COUNTIF(C4:C13,"Working")'
    ws["H2"].font = font(size=1, color=C_WHITE)

    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    headers = ["Staff Name", "Role", *days, "Total Hours"]
    for i, h in enumerate(headers, 2):
        style_header(ws, 3, i, h)

    staff = [
        ("Ahmed Ali",    "Manager"),
        ("Sara Mohamed", "Cashier"),
        ("Omar Hassan",  "Chef"),
        ("Layla Khalid", "Barista"),
        ("Yusuf Salem",  "Waiter"),
        ("Nour Ibrahim", "Waiter"),
        ("Kareem Tarek", "Delivery"),
        ("Hana Adel",    "Cashier"),
    ]

    shifts = ["Working", "Off", "Half Day", "Training"]
    dv_shift = DataValidation(type="list",
                               formula1='"Working,Off,Half Day,Training"',
                               allow_blank=True)
    ws.add_data_validation(dv_shift)

    for r, (name, role) in enumerate(staff, 4):
        ws.row_dimensions[r].height = 24
        bg = C_LIGHT if r % 2 == 0 else C_WHITE
        style_cell(ws, r, 2, name, bold=True, bg=bg, h="left")
        style_cell(ws, r, 3, role, bg=bg)

        for c in range(4, 11):
            ic = input_cell(ws, r, c, "Working" if c < 9 else "Off")
            dv_shift.add(ic)

        # Total hours (Working=8h, Half=4h)
        hours = ws.cell(row=r, column=11,
            value=f'=COUNTIF(D{r}:J{r},"Working")*8+COUNTIF(D{r}:J{r},"Half Day")*4')
        hours.fill = fill(bg)
        hours.font = font(bold=True, color=C_BLUE)
        hours.alignment = align()
        hours.border = border_thin()

    # Summary
    ws.row_dimensions[13].height = 28
    ws.merge_cells("B13:J13")
    style_cell(ws, 13, 2, "WEEKLY TOTALS", bold=True, bg=C_DARK, fg=C_WHITE)
    total_h = ws.cell(row=13, column=11, value="=SUM(K4:K12)")
    total_h.fill = fill(C_DARK)
    total_h.font = font(bold=True, color=C_YELLOW, size=12)
    total_h.alignment = align()


# ══════════════════════════════════════════════════════
#  TAB 5 — REPORTS
# ══════════════════════════════════════════════════════
def build_reports(wb):
    ws = wb.create_sheet("📊 Reports")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_ACCENT

    widths = [2, 25, 18, 18, 18, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 40
    ws.merge_cells("B1:E1")
    t = ws["B1"]
    t.value = "📊  AUTOMATED REPORTS"
    t.fill = fill(C_ACCENT)
    t.font = font(bold=True, size=18, color=C_WHITE)
    t.alignment = align()

    # ── P&L Summary ──────────────────────────────────
    ws.row_dimensions[3].height = 28
    ws.merge_cells("B3:E3")
    style_header(ws, 3, 2, "💰  PROFIT & LOSS SUMMARY (WEEKLY)", bg=C_PRIMARY)

    pl_items = [
        ("Total Revenue",     "='📊 Dashboard'!C19",   C_GREEN),
        ("Total Expenses",    "='📊 Dashboard'!D19",   C_ACCENT),
        ("Gross Profit",      "=C4-C5",                C_BLUE),
        ("Inventory Value",   "='📦 Inventory'!I54",   C_BLUE),
        ("Net Profit",        "=C6-C7",                C_GREEN),
        ("Profit Margin %",   "=IF(C4=0,0,C8/C4*100)", C_ORANGE),
    ]

    for r, (label, formula, color) in enumerate(pl_items, 4):
        ws.row_dimensions[r].height = 26
        bg = C_LIGHT if r % 2 == 0 else C_WHITE
        style_cell(ws, r, 2, label, bold=True, bg=bg, h="left")

        cell = ws.cell(row=r, column=3, value=formula)
        cell.fill = fill(bg)
        cell.font = font(bold=True, color=color, size=13)
        cell.alignment = align()
        cell.border = border_thin()
        if label.endswith("%"):
            cell.number_format = '0.00"%"'
        else:
            cell.number_format = '"$"#,##0.00'

        for c in [4, 5]:
            style_cell(ws, r, c, bg=bg)

    # ── Performance Indicators ───────────────────────
    ws.row_dimensions[11].height = 28
    ws.merge_cells("B11:E11")
    style_header(ws, 11, 2, "📈  PERFORMANCE INDICATORS", bg=C_PRIMARY)

    indicators = [
        ("Best Day Sales",    "=MAX('📊 Dashboard'!C12:C18)"),
        ("Worst Day Sales",   "=MIN(IF('📊 Dashboard'!C12:C18>0,'📊 Dashboard'!C12:C18))"),
        ("Avg Daily Sales",   "=AVERAGE('📊 Dashboard'!C12:C18)"),
        ("Total Orders",      "=SUM('📊 Dashboard'!F12:F18)"),
        ("Avg Order Value",   "=IFERROR(C14/C15,0)"),
        ("Low Stock Items",   "='📦 Inventory'!H2"),
    ]

    for r, (label, formula) in enumerate(indicators, 12):
        ws.row_dimensions[r].height = 24
        bg = C_LIGHT if r % 2 == 0 else C_WHITE
        style_cell(ws, r, 2, label, bold=True, bg=bg, h="left")
        cell = ws.cell(row=r, column=3, value=formula)
        cell.fill = fill(bg)
        cell.font = font(bold=True, color=C_PRIMARY, size=12)
        cell.alignment = align()
        cell.border = border_thin()
        cell.number_format = '"$"#,##0.00'
        for c in [4, 5]:
            style_cell(ws, r, c, bg=bg)

    # ── Quick Tips ───────────────────────────────────
    ws.row_dimensions[19].height = 20
    ws.merge_cells("B19:E19")
    style_header(ws, 19, 2, "💡  QUICK TIPS", bg=C_DARK)

    tips = [
        "✅ Update Daily Sales tab every day for accurate reports",
        "✅ Check Inventory tab weekly to avoid running out of stock",
        "✅ Review this Reports tab every Monday for weekly insights",
        "✅ Yellow cells = input | Green cells = auto-calculated",
        "✅ Share this file with your team via Google Sheets",
    ]
    for r, tip in enumerate(tips, 20):
        ws.row_dimensions[r].height = 22
        bg = "FFF9C4" if r % 2 == 0 else "FFFDE7"
        ws.merge_cells(f"B{r}:E{r}")
        cell = ws[f"B{r}"]
        cell.value = tip
        cell.fill = fill(bg)
        cell.font = font(size=10, color="5D4037")
        cell.alignment = align(h="left")
        cell.border = border_thin()


# ══════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    build_dashboard(wb)
    build_inventory(wb)
    build_daily_sales(wb)
    build_staff(wb)
    build_reports(wb)

    output = "/home/user/smc-bot/Restaurant_Cafe_Manager.xlsx"
    wb.save(output)
    print(f"✅ File created: {output}")
    import os
    size = os.path.getsize(output)
    print(f"📦 Size: {size/1024:.1f} KB")

if __name__ == "__main__":
    main()
