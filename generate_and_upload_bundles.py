"""
NasriTools - Generate Excel Templates + Upload to Bundle Listings
1. Creates all 10 professional Excel templates using openpyxl
2. Uploads the correct files to each of the 5 bundle Etsy listings
3. Publishes all 5 bundles
Run: pip install openpyxl requests && python generate_and_upload_bundles.py
"""
import json, os, time, requests
from pathlib import Path
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

CLIENT_ID  = "pluc0garrgcjzhim0hawxf0k"
SECRET     = "hc89hlqkd6"
SHOP_ID    = 66526082
TOKEN_FILE = Path(os.path.expanduser("~")) / "etsy_token.json"
OUT_DIR    = Path(os.path.expanduser("~")) / "nasri_templates"
OUT_DIR.mkdir(exist_ok=True)

BUNDLE_LISTINGS = {
    "finance_bundle":  4524283814,
    "health_bundle":   4524276503,
    "planner_bundle":  4524276527,
    "business_bundle": 4524276553,
    "ultimate_bundle": 4524283886,
}

BUNDLES = {
    "finance_bundle":  ["budget_tracker", "freelancer_invoice_tracker", "goals_planner"],
    "health_bundle":   ["workout_tracker", "meal_planner", "habit_tracker"],
    "planner_bundle":  ["weekly_planner", "student_planner", "goals_planner"],
    "business_bundle": ["content_creator_planner", "freelancer_invoice_tracker", "goals_planner"],
    "ultimate_bundle": [
        "budget_tracker", "habit_tracker", "meal_planner", "wedding_planner",
        "workout_tracker", "content_creator_planner", "freelancer_invoice_tracker",
        "student_planner", "goals_planner", "weekly_planner",
    ],
}

# ── Style helpers ──────────────────────────────────────────────────────────────

def hdr(ws, row, col, val, bg="1F6B3B", fg="FFFFFF", sz=12, bold=True, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, size=sz, color=fg, name="Calibri")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    return c

def cell(ws, row, col, val="", fmt=None, bold=False, color="000000", align="left", bg=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, size=11, color=color, name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

def border_range(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="CCCCCC")
    b = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(min_row, max_row + 1):
        for c2 in range(min_col, max_col + 1):
            ws.cell(row=r, column=c2).border = b

def alt_row(ws, row, cols, light="F2F9F5"):
    for c2 in range(1, cols + 1):
        ws.cell(row=row, column=c2).fill = PatternFill("solid", fgColor=light)

# ── Template builders ──────────────────────────────────────────────────────────

def make_budget_tracker():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Budget Tracker"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 5
    for col, w in [("B",18),("C",22),("D",15),("E",15),("F",15),("G",18)]:
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 40
    ws.merge_cells("B2:G2")
    t = ws["B2"]; t.value = "💰 Budget Tracker — NasriTools"
    t.font = Font(bold=True, size=18, color="FFFFFF", name="Calibri")
    t.fill = PatternFill("solid", fgColor="1F6B3B")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 8
    # Summary block
    for r,lbl,col in [(4,"Monthly Income","D"),(4,"Total Expenses","E"),(4,"Balance","F")]:
        pass
    hdr(ws,4,2,"Month"); hdr(ws,4,3,"Category"); hdr(ws,4,4,"Description")
    hdr(ws,4,5,"Amount (€)"); hdr(ws,4,6,"Type"); hdr(ws,4,7,"Notes")
    months = ["January","February","March","April","May","June",
              "July","August","September","October","November","December"]
    cats_inc = ["Salary","Freelance","Business","Investment","Other Income"]
    cats_exp = ["Rent/Mortgage","Groceries","Transport","Utilities","Entertainment",
                "Health","Education","Savings","Insurance","Other Expense"]
    row = 5
    for i, m in enumerate(months):
        for cat in cats_inc[:2]:
            cell(ws,row,2,m,align="center"); cell(ws,row,3,cat)
            cell(ws,row,4,""); cell(ws,row,5,0,"€#,##0.00")
            cell(ws,row,6,"Income",color="1F6B3B",bold=True,align="center")
            cell(ws,row,7,"")
            if row % 2 == 0: alt_row(ws,row,7)
            row += 1
        for cat in cats_exp[:3]:
            cell(ws,row,2,m,align="center"); cell(ws,row,3,cat)
            cell(ws,row,4,""); cell(ws,row,5,0,"€#,##0.00")
            cell(ws,row,6,"Expense",color="C0392B",bold=True,align="center")
            cell(ws,row,7,"")
            if row % 2 == 0: alt_row(ws,row,7)
            row += 1
    border_range(ws,4,row-1,2,7)
    # Summary sheet
    ws2 = wb.create_sheet("Monthly Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 4
    for col,w in [("B",20),("C",18),("D",18),("E",18)]:
        ws2.column_dimensions[col].width = w
    ws2.merge_cells("B2:E2")
    t2 = ws2["B2"]; t2.value = "Monthly Summary — NasriTools"
    t2.font = Font(bold=True,size=16,color="FFFFFF"); t2.fill=PatternFill("solid",fgColor="1F6B3B")
    t2.alignment = Alignment(horizontal="center",vertical="center")
    hdr(ws2,4,2,"Month"); hdr(ws2,4,3,"Total Income"); hdr(ws2,4,4,"Total Expenses"); hdr(ws2,4,5,"Balance")
    for i,m in enumerate(months,5):
        cell(ws2,i,2,m,align="center")
        ws2.cell(i,3).number_format="€#,##0.00"
        ws2.cell(i,4).number_format="€#,##0.00"
        ws2.cell(i,5).number_format="€#,##0.00"
        if i%2==0: alt_row(ws2,i,5,"F2F9F5")
    border_range(ws2,4,16,2,5)
    wb.active = ws
    return wb


def make_habit_tracker():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Habit Tracker"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 26
    for i in range(3, 35):
        ws.column_dimensions[get_column_letter(i)].width = 4.5
    ws.column_dimensions[get_column_letter(35)].width = 10
    ws.row_dimensions[2].height = 40
    ws.merge_cells("B2:AH2")
    t = ws["B2"]; t.value = "✅ Habit Tracker — NasriTools"
    t.font=Font(bold=True,size=18,color="FFFFFF"); t.fill=PatternFill("solid",fgColor="1F6B3B")
    t.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[4].height = 28
    hdr(ws,4,2,"Habit / Routine",sz=11)
    for d in range(1,32):
        hdr(ws,4,d+2,str(d),sz=9)
    hdr(ws,4,34,"✓ Days",sz=9)
    habits = [
        "Wake up early (before 7am)","Exercise / Workout","Drink 2L water",
        "Read 20 min","Meditate 10 min","No social media before 9am",
        "Healthy breakfast","Journal / Gratitude","Study / Learning",
        "Walk 10,000 steps","Sleep before midnight","Cold shower",
        "No junk food","Cook at home","Practice language",
        "Call family / friends","Digital detox 1hr","Review daily goals",
        "Stretch / Yoga","Vitamins / Supplements","No alcohol","Save money today",
        "Work on side project","No screens after 10pm","Floss & skincare",
        "Clean workspace","Inbox zero","Plan tomorrow","Creative work","Acts of kindness",
    ]
    for i,h in enumerate(habits,5):
        ws.row_dimensions[i].height = 22
        cell(ws,i,2,h,bold=(i%5==0))
        for d in range(1,32):
            c=ws.cell(row=i,column=d+2)
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.fill=PatternFill("solid",fgColor=("F0F0F0" if i%2==0 else "FFFFFF"))
        ws.cell(i,34).number_format="0"
        ws.cell(i,34).alignment=Alignment(horizontal="center",vertical="center")
        ws.cell(i,34).font=Font(bold=True,color="1F6B3B")
        if i%2==0: alt_row(ws,i,34,"F2F9F5")
    border_range(ws,4,34,2,34)
    return wb


def make_meal_planner():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Meal Planner"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    for col,w in [("B",14),("C",22),("D",22),("E",22),("F",22),("G",18)]:
        ws.column_dimensions[col].width = w
    ws.row_dimensions[2].height = 40
    ws.merge_cells("B2:G2")
    t=ws["B2"]; t.value="🥗 Meal Planner — NasriTools"
    t.font=Font(bold=True,size=18,color="FFFFFF"); t.fill=PatternFill("solid",fgColor="2E86AB")
    t.alignment=Alignment(horizontal="center",vertical="center")
    hdr(ws,4,2,"Day",bg="2E86AB"); hdr(ws,4,3,"Breakfast",bg="2E86AB")
    hdr(ws,4,4,"Lunch",bg="2E86AB"); hdr(ws,4,5,"Dinner",bg="2E86AB")
    hdr(ws,4,6,"Snacks",bg="2E86AB"); hdr(ws,4,7,"Calories",bg="2E86AB")
    days=["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    for i,d in enumerate(days,5):
        ws.row_dimensions[i].height = 28
        cell(ws,i,2,d,bold=True,align="center")
        for c2 in range(3,7): ws.cell(i,c2).alignment=Alignment(wrap_text=True,vertical="top")
        ws.cell(i,7).number_format="0"
        ws.cell(i,7).alignment=Alignment(horizontal="center",vertical="center")
        if i%2==0: alt_row(ws,i,7,"EAF4FB")
    border_range(ws,4,11,2,7)
    # Grocery list
    ws2=wb.create_sheet("Grocery List")
    ws2.sheet_view.showGridLines=False
    ws2.column_dimensions["A"].width=4
    for col,w in [("B",28),("C",14),("D",14),("E",16)]:
        ws2.column_dimensions[col].width=w
    ws2.merge_cells("B2:E2")
    t2=ws2["B2"]; t2.value="🛒 Grocery List — NasriTools"
    t2.font=Font(bold=True,size=16,color="FFFFFF"); t2.fill=PatternFill("solid",fgColor="2E86AB")
    t2.alignment=Alignment(horizontal="center",vertical="center")
    hdr(ws2,4,2,"Item",bg="2E86AB"); hdr(ws2,4,3,"Qty",bg="2E86AB")
    hdr(ws2,4,4,"Unit",bg="2E86AB"); hdr(ws2,4,5,"Est. Cost",bg="2E86AB")
    for i in range(5,35):
        ws2.row_dimensions[i].height=22
        ws2.cell(i,5).number_format="€#,##0.00"
        if i%2==0: alt_row(ws2,i,5,"EAF4FB")
    border_range(ws2,4,34,2,5)
    return wb


def make_wedding_planner():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Wedding Budget"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    for col,w in [("B",22),("C",24),("D",15),("E",15),("F",15),("G",16),("H",18)]:
        ws.column_dimensions[col].width = w
    ws.row_dimensions[2].height = 40
    ws.merge_cells("B2:H2")
    t=ws["B2"]; t.value="💍 Wedding Planner — NasriTools"
    t.font=Font(bold=True,size=18,color="FFFFFF"); t.fill=PatternFill("solid",fgColor="8E44AD")
    t.alignment=Alignment(horizontal="center",vertical="center")
    hdr(ws,4,2,"Category",bg="8E44AD"); hdr(ws,4,3,"Vendor/Item",bg="8E44AD")
    hdr(ws,4,4,"Budgeted",bg="8E44AD"); hdr(ws,4,5,"Actual",bg="8E44AD")
    hdr(ws,4,6,"Deposit Paid",bg="8E44AD"); hdr(ws,4,7,"Status",bg="8E44AD")
    hdr(ws,4,8,"Notes",bg="8E44AD")
    cats=[("Venue","Ceremony & Reception hall"),("Catering","Food & drinks"),
          ("Photography","Photographer & videographer"),("Music","Band or DJ"),
          ("Flowers","Bouquets & decorations"),("Dress","Wedding dress & alterations"),
          ("Suit","Groom attire"),("Cake","Wedding cake"),
          ("Transport","Limo / cars"),("Honeymoon","Flights & hotel"),
          ("Invitations","Cards & postage"),("Hair & Makeup","Bride & bridal party"),
          ("Rings","Wedding bands"),("Entertainment","Photo booth & extras"),
          ("Officiant","Ceremony officiant"),("Gifts","Favors & bridal party gifts"),
          ("Other","Miscellaneous")]
    for i,(cat,item) in enumerate(cats,5):
        ws.row_dimensions[i].height=24
        cell(ws,i,2,cat,bold=True); cell(ws,i,3,item)
        for c2 in [4,5,6]: ws.cell(i,c2).number_format="€#,##0.00"
        ws.cell(i,7).alignment=Alignment(horizontal="center",vertical="center")
        if i%2==0: alt_row(ws,i,8,"F5EEF8")
    border_range(ws,4,21,2,8)
    # Guest list
    ws2=wb.create_sheet("Guest List")
    ws2.sheet_view.showGridLines=False
    ws2.column_dimensions["A"].width=4
    for col,w in [("B",5),("C",22),("D",22),("E",14),("F",16),("G",16),("H",16)]:
        ws2.column_dimensions[col].width=w
    ws2.merge_cells("B2:H2")
    t2=ws2["B2"]; t2.value="Guest List"
    t2.font=Font(bold=True,size=16,color="FFFFFF"); t2.fill=PatternFill("solid",fgColor="8E44AD")
    t2.alignment=Alignment(horizontal="center",vertical="center")
    hdr(ws2,4,2,"#",bg="8E44AD"); hdr(ws2,4,3,"First Name",bg="8E44AD")
    hdr(ws2,4,4,"Last Name",bg="8E44AD"); hdr(ws2,4,5,"Side",bg="8E44AD")
    hdr(ws2,4,6,"Invited",bg="8E44AD"); hdr(ws2,4,7,"RSVP",bg="8E44AD")
    hdr(ws2,4,8,"Dietary",bg="8E44AD")
    for i in range(5,105):
        ws2.row_dimensions[i].height=20
        ws2.cell(i,2).value=i-4
        ws2.cell(i,2).alignment=Alignment(horizontal="center",vertical="center")
        if i%2==0: alt_row(ws2,i,8,"F5EEF8")
    border_range(ws2,4,104,2,8)
    return wb


def make_workout_tracker():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Workout Log"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    for col,w in [("B",8),("C",24),("D",18),("E",10),("F",10),("G",12),("H",12),("I",18)]:
        ws.column_dimensions[col].width = w
    ws.row_dimensions[2].height = 40
    ws.merge_cells("B2:I2")
    t=ws["B2"]; t.value="💪 Workout Tracker — NasriTools"
    t.font=Font(bold=True,size=18,color="FFFFFF"); t.fill=PatternFill("solid",fgColor="C0392B")
    t.alignment=Alignment(horizontal="center",vertical="center")
    hdr(ws,4,2,"Date",bg="C0392B"); hdr(ws,4,3,"Exercise",bg="C0392B")
    hdr(ws,4,4,"Muscle Group",bg="C0392B"); hdr(ws,4,5,"Sets",bg="C0392B")
    hdr(ws,4,6,"Reps",bg="C0392B"); hdr(ws,4,7,"Weight (kg)",bg="C0392B")
    hdr(ws,4,8,"PR?",bg="C0392B"); hdr(ws,4,9,"Notes",bg="C0392B")
    exercises=[
        ("Bench Press","Chest"),("Push-ups","Chest"),("Incline DB Press","Chest"),
        ("Pull-ups","Back"),("Deadlift","Back"),("Lat Pulldown","Back"),
        ("Squat","Legs"),("Leg Press","Legs"),("Romanian Deadlift","Legs"),
        ("Overhead Press","Shoulders"),("Lateral Raises","Shoulders"),
        ("Bicep Curl","Arms"),("Tricep Dips","Arms"),
        ("Plank","Core"),("Crunches","Core"),
        ("Running","Cardio"),("Cycling","Cardio"),("Jump Rope","Cardio"),
    ]
    for i,(_ex,mg) in enumerate(exercises*3,5):
        ws.row_dimensions[i].height=22
        ws.cell(i,2).number_format="DD/MM/YYYY"
        cell(ws,i,3,_ex); cell(ws,i,4,mg,align="center")
        for c2 in [5,6]: ws.cell(i,c2).number_format="0"; ws.cell(i,c2).alignment=Alignment(horizontal="center",vertical="center")
        ws.cell(i,7).number_format="0.0"; ws.cell(i,7).alignment=Alignment(horizontal="center",vertical="center")
        ws.cell(i,8).alignment=Alignment(horizontal="center",vertical="center")
        if i%2==0: alt_row(ws,i,9,"FDF0EE")
    border_range(ws,4,4+len(exercises)*3,2,9)
    return wb


def make_content_creator_planner():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Content Calendar"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    for col,w in [("B",12),("C",14),("D",28),("E",18),("F",16),("G",14),("H",16),("I",16)]:
        ws.column_dimensions[col].width = w
    ws.row_dimensions[2].height = 40
    ws.merge_cells("B2:I2")
    t=ws["B2"]; t.value="📱 Content Creator Planner — NasriTools"
    t.font=Font(bold=True,size=18,color="FFFFFF"); t.fill=PatternFill("solid",fgColor="E67E22")
    t.alignment=Alignment(horizontal="center",vertical="center")
    hdr(ws,4,2,"Date",bg="E67E22"); hdr(ws,4,3,"Platform",bg="E67E22")
    hdr(ws,4,4,"Content Idea / Caption",bg="E67E22"); hdr(ws,4,5,"Format",bg="E67E22")
    hdr(ws,4,6,"Status",bg="E67E22"); hdr(ws,4,7,"Likes",bg="E67E22")
    hdr(ws,4,8,"Comments",bg="E67E22"); hdr(ws,4,9,"Reach",bg="E67E22")
    platforms=["Instagram","YouTube","TikTok","Twitter/X","LinkedIn","Pinterest","Facebook","Blog"]
    formats=["Reel","Post","Story","Short","Video","Carousel","Thread","Article"]
    statuses=["Idea","Scripted","Recorded","Edited","Scheduled","Posted"]
    for i in range(5,65):
        ws.row_dimensions[i].height=24
        ws.cell(i,2).number_format="DD/MM/YYYY"
        for c2 in [7,8,9]: ws.cell(i,c2).number_format="#,##0"; ws.cell(i,c2).alignment=Alignment(horizontal="center",vertical="center")
        ws.cell(i,4).alignment=Alignment(wrap_text=True,vertical="top")
        if i%2==0: alt_row(ws,i,9,"FEF5EC")
    border_range(ws,4,64,2,9)
    return wb


def make_freelancer_invoice_tracker():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Invoice Tracker"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    for col,w in [("B",14),("C",22),("D",24),("E",15),("F",15),("G",15),("H",14),("I",16)]:
        ws.column_dimensions[col].width = w
    ws.row_dimensions[2].height = 40
    ws.merge_cells("B2:I2")
    t=ws["B2"]; t.value="📄 Freelancer Invoice Tracker — NasriTools"
    t.font=Font(bold=True,size=18,color="FFFFFF"); t.fill=PatternFill("solid",fgColor="2980B9")
    t.alignment=Alignment(horizontal="center",vertical="center")
    hdr(ws,4,2,"Invoice #",bg="2980B9"); hdr(ws,4,3,"Client Name",bg="2980B9")
    hdr(ws,4,4,"Project / Service",bg="2980B9"); hdr(ws,4,5,"Invoice Date",bg="2980B9")
    hdr(ws,4,6,"Due Date",bg="2980B9"); hdr(ws,4,7,"Amount (€)",bg="2980B9")
    hdr(ws,4,8,"Status",bg="2980B9"); hdr(ws,4,9,"Notes",bg="2980B9")
    for i in range(5,55):
        ws.row_dimensions[i].height=24
        ws.cell(i,2).value=f"INV-{i-4:03d}"
        ws.cell(i,2).alignment=Alignment(horizontal="center",vertical="center")
        for c2 in [5,6]: ws.cell(i,c2).number_format="DD/MM/YYYY"
        ws.cell(i,7).number_format="€#,##0.00"; ws.cell(i,7).alignment=Alignment(horizontal="right",vertical="center")
        ws.cell(i,8).alignment=Alignment(horizontal="center",vertical="center")
        if i%2==0: alt_row(ws,i,9,"EBF5FB")
    border_range(ws,4,54,2,9)
    # Client DB
    ws2=wb.create_sheet("Client Database")
    ws2.sheet_view.showGridLines=False
    ws2.column_dimensions["A"].width=4
    for col,w in [("B",22),("C",28),("D",20),("E",18),("F",16),("G",20)]:
        ws2.column_dimensions[col].width=w
    ws2.merge_cells("B2:G2")
    t2=ws2["B2"]; t2.value="Client Database"
    t2.font=Font(bold=True,size=16,color="FFFFFF"); t2.fill=PatternFill("solid",fgColor="2980B9")
    t2.alignment=Alignment(horizontal="center",vertical="center")
    hdr(ws2,4,2,"Client Name",bg="2980B9"); hdr(ws2,4,3,"Email",bg="2980B9")
    hdr(ws2,4,4,"Phone",bg="2980B9"); hdr(ws2,4,5,"Country",bg="2980B9")
    hdr(ws2,4,6,"Hourly Rate",bg="2980B9"); hdr(ws2,4,7,"Notes",bg="2980B9")
    for i in range(5,55):
        ws2.row_dimensions[i].height=22
        ws2.cell(i,6).number_format="€#,##0.00"
        if i%2==0: alt_row(ws2,i,7,"EBF5FB")
    border_range(ws2,4,54,2,7)
    return wb


def make_student_planner():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Assignment Tracker"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    for col,w in [("B",20),("C",28),("D",12),("E",14),("F",14),("G",16),("H",16)]:
        ws.column_dimensions[col].width = w
    ws.row_dimensions[2].height = 40
    ws.merge_cells("B2:H2")
    t=ws["B2"]; t.value="🎓 Student Planner — NasriTools"
    t.font=Font(bold=True,size=18,color="FFFFFF"); t.fill=PatternFill("solid",fgColor="1A5276")
    t.alignment=Alignment(horizontal="center",vertical="center")
    hdr(ws,4,2,"Subject",bg="1A5276"); hdr(ws,4,3,"Assignment / Task",bg="1A5276")
    hdr(ws,4,4,"Due Date",bg="1A5276"); hdr(ws,4,5,"Priority",bg="1A5276")
    hdr(ws,4,6,"Status",bg="1A5276"); hdr(ws,4,7,"Grade",bg="1A5276")
    hdr(ws,4,8,"Notes",bg="1A5276")
    subjects=["Mathematics","Physics","Chemistry","Biology","History","English Literature",
              "Economics","Computer Science","Psychology","Philosophy","Geography","Art"]
    for i,s in enumerate(subjects*4,5):
        ws.row_dimensions[i].height=24
        cell(ws,i,2,s,bold=True); ws.cell(i,4).number_format="DD/MM/YYYY"
        ws.cell(i,5).alignment=Alignment(horizontal="center",vertical="center")
        ws.cell(i,6).alignment=Alignment(horizontal="center",vertical="center")
        ws.cell(i,7).alignment=Alignment(horizontal="center",vertical="center")
        if i%2==0: alt_row(ws,i,8,"EBF5FB")
    border_range(ws,4,4+len(subjects)*4,2,8)
    # GPA tracker
    ws2=wb.create_sheet("GPA Tracker")
    ws2.sheet_view.showGridLines=False
    ws2.column_dimensions["A"].width=4
    for col,w in [("B",22),("C",14),("D",14),("E",14),("F",14)]:
        ws2.column_dimensions[col].width=w
    ws2.merge_cells("B2:F2")
    t2=ws2["B2"]; t2.value="GPA Tracker"
    t2.font=Font(bold=True,size=16,color="FFFFFF"); t2.fill=PatternFill("solid",fgColor="1A5276")
    t2.alignment=Alignment(horizontal="center",vertical="center")
    hdr(ws2,4,2,"Subject",bg="1A5276"); hdr(ws2,4,3,"Credits",bg="1A5276")
    hdr(ws2,4,4,"Grade %",bg="1A5276"); hdr(ws2,4,5,"Grade Letter",bg="1A5276")
    hdr(ws2,4,6,"Grade Points",bg="1A5276")
    for i in range(5,25):
        ws2.row_dimensions[i].height=22
        for c2 in [3,4,5,6]: ws2.cell(i,c2).alignment=Alignment(horizontal="center",vertical="center")
        if i%2==0: alt_row(ws2,i,6,"EBF5FB")
    border_range(ws2,4,24,2,6)
    return wb


def make_goals_planner():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Annual Goals"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    for col,w in [("B",6),("C",26),("D",18),("E",16),("F",16),("G",14),("H",18)]:
        ws.column_dimensions[col].width = w
    ws.row_dimensions[2].height = 40
    ws.merge_cells("B2:H2")
    t=ws["B2"]; t.value="🎯 Goals Planner — NasriTools"
    t.font=Font(bold=True,size=18,color="FFFFFF"); t.fill=PatternFill("solid",fgColor="117A65")
    t.alignment=Alignment(horizontal="center",vertical="center")
    hdr(ws,4,2,"#",bg="117A65"); hdr(ws,4,3,"Goal",bg="117A65")
    hdr(ws,4,4,"Category",bg="117A65"); hdr(ws,4,5,"Target Date",bg="117A65")
    hdr(ws,4,6,"Why (Motivation)",bg="117A65"); hdr(ws,4,7,"Progress %",bg="117A65")
    hdr(ws,4,8,"Status",bg="117A65")
    categories=["Health","Finance","Career","Education","Relationships",
                "Personal Growth","Travel","Business","Creative","Fitness"]
    for i in range(5,17):
        ws.row_dimensions[i].height=36
        ws.cell(i,2).value=i-4; ws.cell(i,2).alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.cell(i,3).alignment=Alignment(wrap_text=True,vertical="top")
        ws.cell(i,4).value=categories[(i-5)%len(categories)]
        ws.cell(i,4).alignment=Alignment(horizontal="center",vertical="center")
        ws.cell(i,5).number_format="DD/MM/YYYY"
        ws.cell(i,6).alignment=Alignment(wrap_text=True,vertical="top")
        ws.cell(i,7).number_format="0%"; ws.cell(i,7).alignment=Alignment(horizontal="center",vertical="center")
        ws.cell(i,8).alignment=Alignment(horizontal="center",vertical="center")
        if i%2==0: alt_row(ws,i,8,"E8F8F5")
    border_range(ws,4,16,2,8)
    # 90-day plan
    ws2=wb.create_sheet("90-Day Action Plan")
    ws2.sheet_view.showGridLines=False
    ws2.column_dimensions["A"].width=4
    for col,w in [("B",8),("C",26),("D",18),("E",16),("F",14),("G",14)]:
        ws2.column_dimensions[col].width=w
    ws2.merge_cells("B2:G2")
    t2=ws2["B2"]; t2.value="90-Day Action Plan"
    t2.font=Font(bold=True,size=16,color="FFFFFF"); t2.fill=PatternFill("solid",fgColor="117A65")
    t2.alignment=Alignment(horizontal="center",vertical="center")
    hdr(ws2,4,2,"Week",bg="117A65"); hdr(ws2,4,3,"Action Step",bg="117A65")
    hdr(ws2,4,4,"Related Goal",bg="117A65"); hdr(ws2,4,5,"Deadline",bg="117A65")
    hdr(ws2,4,6,"Done?",bg="117A65"); hdr(ws2,4,7,"Notes",bg="117A65")
    for i in range(5,95):
        ws2.row_dimensions[i].height=22
        ws2.cell(i,2).value=f"Week {((i-5)//2)+1}"
        ws2.cell(i,2).alignment=Alignment(horizontal="center",vertical="center")
        ws2.cell(i,5).number_format="DD/MM/YYYY"
        ws2.cell(i,6).alignment=Alignment(horizontal="center",vertical="center")
        if i%2==0: alt_row(ws2,i,7,"E8F8F5")
    border_range(ws2,4,94,2,7)
    return wb


def make_weekly_planner():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Weekly Planner"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 14
    for col in ["C","D","E","F","G","H","I"]:
        ws.column_dimensions[col].width = 20
    ws.row_dimensions[2].height = 40
    ws.merge_cells("B2:I2")
    t=ws["B2"]; t.value="📅 Weekly Planner — NasriTools"
    t.font=Font(bold=True,size=18,color="FFFFFF"); t.fill=PatternFill("solid",fgColor="6C3483")
    t.alignment=Alignment(horizontal="center",vertical="center")
    hdr(ws,4,2,"Time Slot",bg="6C3483")
    days=["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    for j,d in enumerate(days,3):
        hdr(ws,4,j,d,bg="6C3483")
    times=["6:00","7:00","8:00","9:00","10:00","11:00","12:00",
           "13:00","14:00","15:00","16:00","17:00","18:00","19:00",
           "20:00","21:00","22:00"]
    for i,t2 in enumerate(times,5):
        ws.row_dimensions[i].height=28
        cell(ws,i,2,t2,bold=True,align="center")
        for j in range(3,10):
            ws.cell(i,j).alignment=Alignment(wrap_text=True,vertical="top")
            if i%2==0:
                ws.cell(i,j).fill=PatternFill("solid",fgColor="F4ECF7")
    border_range(ws,4,4+len(times),2,9)
    # Top priorities
    ws2=wb.create_sheet("Weekly Goals")
    ws2.sheet_view.showGridLines=False
    ws2.column_dimensions["A"].width=4
    for col,w in [("B",6),("C",30),("D",18),("E",16),("F",14)]:
        ws2.column_dimensions[col].width=w
    ws2.merge_cells("B2:F2")
    t3=ws2["B2"]; t3.value="Weekly Goals & Priorities"
    t3.font=Font(bold=True,size=16,color="FFFFFF"); t3.fill=PatternFill("solid",fgColor="6C3483")
    t3.alignment=Alignment(horizontal="center",vertical="center")
    hdr(ws2,4,2,"#",bg="6C3483"); hdr(ws2,4,3,"Goal / Task",bg="6C3483")
    hdr(ws2,4,4,"Category",bg="6C3483"); hdr(ws2,4,5,"Priority",bg="6C3483")
    hdr(ws2,4,6,"Done?",bg="6C3483")
    for i in range(5,20):
        ws2.row_dimensions[i].height=24
        ws2.cell(i,2).value=i-4; ws2.cell(i,2).alignment=Alignment(horizontal="center",vertical="center")
        ws2.cell(i,5).alignment=Alignment(horizontal="center",vertical="center")
        ws2.cell(i,6).alignment=Alignment(horizontal="center",vertical="center")
        if i%2==0: alt_row(ws2,i,6,"F4ECF7")
    border_range(ws2,4,19,2,6)
    return wb


# ── Etsy helpers ───────────────────────────────────────────────────────────────

def get_token():
    t = json.loads(TOKEN_FILE.read_text())
    if time.time() >= t.get("expires_at", 0):
        r = requests.post("https://api.etsy.com/v3/public/oauth/token", data={
            "grant_type": "refresh_token", "client_id": CLIENT_ID,
            "refresh_token": t["refresh_token"],
        })
        r.raise_for_status()
        t = r.json()
        t["expires_at"] = time.time() + t.get("expires_in", 3600) - 60
        TOKEN_FILE.write_text(json.dumps(t, indent=2))
    return t


def auth_headers(token):
    return {"Authorization": "Bearer " + token["access_token"],
            "x-api-key": CLIENT_ID + ":" + SECRET}


def upload_xlsx_to_listing(token, listing_id, filepath):
    with open(filepath, "rb") as f:
        r = requests.post(
            f"https://api.etsy.com/v3/application/shops/{SHOP_ID}/listings/{listing_id}/files",
            headers=auth_headers(token),
            files={"file": (filepath.name, f,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"name": filepath.name, "rank": 1},
            timeout=60,
        )
    return r


def publish_listing(token, listing_id):
    r = requests.patch(
        f"https://api.etsy.com/v3/application/shops/{SHOP_ID}/listings/{listing_id}",
        headers={**auth_headers(token), "Content-Type": "application/json"},
        json={"state": "active"},
        timeout=30,
    )
    return r


# ── Main ───────────────────────────────────────────────────────────────────────

GENERATORS = {
    "budget_tracker":             make_budget_tracker,
    "habit_tracker":              make_habit_tracker,
    "meal_planner":               make_meal_planner,
    "wedding_planner":            make_wedding_planner,
    "workout_tracker":            make_workout_tracker,
    "content_creator_planner":    make_content_creator_planner,
    "freelancer_invoice_tracker": make_freelancer_invoice_tracker,
    "student_planner":            make_student_planner,
    "goals_planner":              make_goals_planner,
    "weekly_planner":             make_weekly_planner,
}

LABELS = {
    "budget_tracker":             "Budget Tracker",
    "habit_tracker":              "Habit Tracker",
    "meal_planner":               "Meal Planner",
    "wedding_planner":            "Wedding Planner",
    "workout_tracker":            "Workout Tracker",
    "content_creator_planner":    "Content Creator Planner",
    "freelancer_invoice_tracker": "Freelancer Invoice Tracker",
    "student_planner":            "Student Planner",
    "goals_planner":              "Goals Planner",
    "weekly_planner":             "Weekly Planner",
}


def main():
    print(f"\n{'='*65}")
    print(f"  NasriTools - Generate Templates + Upload Bundles")
    print(f"{'='*65}\n")

    token = get_token()

    # Step 1: generate all 10 xlsx files
    print("  Step 1: Generating 10 Excel templates...\n")
    generated = {}
    for slug, gen_fn in GENERATORS.items():
        path = OUT_DIR / f"{slug}.xlsx"
        if not path.exists():
            wb = gen_fn()
            wb.save(path)
            print(f"  ✓  {LABELS[slug]}  →  {path.name}")
        else:
            print(f"  —  {LABELS[slug]}  (already exists, skipping)")
        generated[slug] = path

    # Step 2: upload to bundle listings
    print(f"\n  Step 2: Uploading files to bundle listings...\n")
    for bkey, slugs in BUNDLES.items():
        lid = BUNDLE_LISTINGS[bkey]
        bname = bkey.replace("_", " ").title()
        print(f"  [{bname}]  listing {lid}")

        for slug in slugs:
            fpath = generated[slug]
            print(f"    Uploading {fpath.name}...", end=" ")
            r = upload_xlsx_to_listing(token, lid, fpath)
            time.sleep(1)
            if r.ok:
                print("✓")
            else:
                print(f"✗  {r.status_code}: {r.text[:80]}")

        # Publish
        print(f"    Publishing listing...", end=" ")
        rp = publish_listing(token, lid)
        time.sleep(1)
        if rp.ok:
            print(f"✓  https://www.etsy.com/listing/{lid}")
        else:
            print(f"✗  {rp.status_code}: {rp.text[:80]}")
        print()

    print(f"{'='*65}")
    print(f"  Done! Templates saved to: {OUT_DIR}")
    print(f"  All 5 bundles published on Etsy.")
    print(f"{'='*65}\n")


if __name__ == "__main__":
    main()
