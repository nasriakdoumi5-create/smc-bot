"""
NasriTools - Create Free Budget Tracker (Lead Magnet)
Creates a simplified FREE version of the Budget Tracker to:
  1. Attract organic traffic and favorites
  2. Signal product quality to Etsy algorithm
  3. Convert free users to paid bundle buyers

Run: python create_free_product.py
"""
import json, os, time, requests, io
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CLIENT_ID  = "pluc0garrgcjzhim0hawxf0k"
SECRET     = "hc89hlqkd6"
SHOP_ID    = 66526082
TOKEN_FILE = Path(os.path.expanduser("~")) / "etsy_token.json"
OUT_DIR    = Path(os.path.expanduser("~")) / "nasri_free"
OUT_DIR.mkdir(exist_ok=True)

BRAND_COLOR = (22, 100, 52)
WHITE       = (255, 255, 255)
GOLD        = (255, 195, 0)
DARK        = (18, 22, 35)
LGRAY       = (240, 242, 248)
GRAY        = (110, 115, 130)
SIZE        = 2000


def get_token():
    t = json.loads(TOKEN_FILE.read_text())
    if time.time() >= t.get("expires_at", 0) - 60:
        r = requests.post("https://api.etsy.com/v3/public/oauth/token", data={
            "grant_type": "refresh_token", "client_id": CLIENT_ID,
            "refresh_token": t["refresh_token"],
        })
        r.raise_for_status()
        t = r.json()
        t["expires_at"] = time.time() + t.get("expires_in", 3600) - 60
        TOKEN_FILE.write_text(json.dumps(t, indent=2))
    return t


def auth_headers(token):
    return {"Authorization": "Bearer " + token["access_token"],
            "x-api-key": CLIENT_ID + ":" + SECRET}


def load_font(size):
    for f in [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
    ]:
        if Path(f).exists():
            try:
                return ImageFont.truetype(f, size)
            except Exception:
                pass
    return ImageFont.load_default()


def load_font_reg(size):
    for f in [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "C:/Windows/Fonts/arial.ttf",
    ]:
        if Path(f).exists():
            try:
                return ImageFont.truetype(f, size)
            except Exception:
                pass
    return ImageFont.load_default()


def draw_centered(draw, text, font, cx, y, fill):
    bbox = draw.textbbox((0, 0), text, font=font)
    w = bbox[2] - bbox[0]
    draw.text((cx - w // 2, y), text, font=font, fill=fill)


def make_free_thumbnail():
    color = BRAND_COLOR
    r, g, b = color

    img  = Image.new("RGB", (SIZE, SIZE), color)
    draw = ImageDraw.Draw(img)

    c2 = (min(r+30, 255), min(g+30, 255), min(b+30, 255))
    draw.ellipse([SIZE-560, -280, SIZE+260, 560], fill=c2)
    draw.ellipse([-200, SIZE-380, 280, SIZE+180], fill=c2)

    # FREE badge
    draw.ellipse([SIZE//2-170, 60, SIZE//2+170, 400], fill=GOLD)
    draw_centered(draw, "FREE", load_font(130), SIZE//2, 120, (50, 30, 0))
    draw_centered(draw, "DOWNLOAD", load_font(42), SIZE//2, 300, (80, 50, 0))

    draw.text((70, 60), "NasriTools", font=load_font(52), fill=(255, 255, 200))

    try:
        draw.text((SIZE//2, 600), "💰", font=load_font(180), fill=WHITE, anchor="mm")
    except Exception:
        pass

    draw_centered(draw, "BUDGET TRACKER", load_font(110), SIZE//2, 740, WHITE)
    draw_centered(draw, "Simple Edition", load_font_reg(52), SIZE//2, 890, (200, 230, 210))

    draw.rectangle([SIZE//2-240, 980, SIZE//2+240, 988], fill=WHITE)

    features = ["Track income & expenses", "Monthly budget overview", "Savings goal tracker"]
    fy = 1010
    for feat in features:
        draw_centered(draw, "✓  " + feat, load_font_reg(44), SIZE//2, fy, (200, 240, 215))
        fy += 65

    draw.rounded_rectangle([SIZE//2-380, 1230, SIZE//2+380, 1340], radius=24, fill=WHITE)
    draw_centered(draw, "🔓 Want the FULL version? See our shop!",
                  load_font_reg(38), SIZE//2, 1258, color)
    draw_centered(draw, "10 templates — starting at €4.99",
                  load_font(34), SIZE//2, 1305, color)

    draw.rectangle([0, SIZE-110, SIZE, SIZE], fill=(r//3, g//3, b//3))
    draw_centered(draw, "100% Free  •  No Email Required  •  Instant Download",
                  load_font_reg(34), SIZE//2, SIZE-72, WHITE)

    return img


def make_free_budget_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Tracker (Free)"

    green = "16643C"
    light_green = "D6F5E4"
    gold  = "FFC300"
    white = "FFFFFF"
    dark  = "1C2233"

    def s(row, col, value, bold=False, bg=None, fg="000000", size=11, align="left"):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, size=size, color=fg, name="Calibri")
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        bd = Side(style="thin", color="DDDDDD")
        c.border = Border(left=bd, right=bd, top=bd, bottom=bd)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.row_dimensions[1].height = 52

    ws.merge_cells("A1:D1")
    s(1, 1, "💰 BUDGET TRACKER — Simple Edition (Free)", bold=True, bg=green, fg=white, size=16, align="center")

    ws.merge_cells("A2:D2")
    s(2, 1, "NasriTools | nasritools.etsy.com | Upgrade to full version for 12-month tracking!", bg=light_green, fg=green, size=10, align="center")

    ws.row_dimensions[4].height = 40
    ws.merge_cells("A4:D4")
    s(4, 1, "📅 MONTHLY BUDGET", bold=True, bg=dark, fg=gold, size=14, align="center")

    ws.row_dimensions[5].height = 32
    for col, header in enumerate(["Category", "Budget (€)", "Actual (€)", "Difference (€)"], 1):
        s(5, col, header, bold=True, bg=green, fg=white, size=12, align="center")

    ws.merge_cells("A6:D6")
    s(6, 1, "INCOME", bold=True, bg=light_green, fg=green, size=12)

    income = [("Salary / Main Income", 2500, 2500), ("Side Income / Freelance", 500, 350), ("Other", 0, 0)]
    for i, (cat, budget, actual) in enumerate(income, 7):
        ws.row_dimensions[i].height = 26
        bg = "EAF7EE" if i % 2 == 0 else white
        s(i, 1, cat, bg=bg)
        s(i, 2, budget, align="center", bg=bg)
        s(i, 3, actual, align="center", bg=bg)
        ws.cell(row=i, column=4).value = f"=C{i}-B{i}"
        ws.cell(row=i, column=4).font = Font(size=11, name="Calibri")
        ws.cell(row=i, column=4).fill = PatternFill("solid", fgColor=bg)
        ws.cell(row=i, column=4).alignment = Alignment(horizontal="center", vertical="center")

    exp_start = 11
    ws.merge_cells(f"A{exp_start}:D{exp_start}")
    s(exp_start, 1, "EXPENSES", bold=True, bg="FFF0E0", fg="CC5500", size=12)

    expenses = [
        ("Housing (Rent / Mortgage)", 800, 800),
        ("Food & Groceries", 300, 265),
        ("Transport", 120, 98),
        ("Utilities (Electric/Internet)", 80, 75),
        ("Healthcare", 50, 0),
        ("Entertainment & Subscriptions", 60, 85),
        ("Savings & Investments", 300, 300),
        ("Other Expenses", 0, 22),
    ]
    exp_end = exp_start
    for i, (cat, budget, actual) in enumerate(expenses, exp_start + 1):
        ws.row_dimensions[i].height = 26
        bg = "FFF8F0" if i % 2 == 0 else white
        s(i, 1, cat, bg=bg)
        s(i, 2, budget, align="center", bg=bg)
        s(i, 3, actual, align="center", bg=bg)
        ws.cell(row=i, column=4).value = f"=C{i}-B{i}"
        ws.cell(row=i, column=4).font = Font(size=11, name="Calibri")
        ws.cell(row=i, column=4).fill = PatternFill("solid", fgColor=bg)
        ws.cell(row=i, column=4).alignment = Alignment(horizontal="center", vertical="center")
        exp_end = i

    sum_row = exp_end + 2
    ws.merge_cells(f"A{sum_row}:D{sum_row}")
    s(sum_row, 1, "💹 MONTHLY SUMMARY", bold=True, bg=dark, fg=gold, size=13, align="center")

    for off, label, b_f, c_f, bg, fg in [
        (1, "Total Income",   f"=SUM(B7:B9)",   f"=SUM(C7:C9)",   green,   white),
        (2, "Total Expenses", f"=SUM(B{exp_start+1}:B{exp_end})", f"=SUM(C{exp_start+1}:C{exp_end})", "CC5500", white),
        (3, "NET SAVINGS",    f"=B{sum_row+1}-B{sum_row+2}", f"=C{sum_row+1}-C{sum_row+2}", gold, dark),
    ]:
        rn = sum_row + off
        ws.row_dimensions[rn].height = 32
        s(rn, 1, label, bold=True, bg=bg, fg=fg, size=12)
        for col_i, formula in [(2, b_f), (3, c_f)]:
            c = ws.cell(row=rn, column=col_i)
            c.value = formula
            c.font  = Font(bold=True, size=12, color=fg, name="Calibri")
            c.fill  = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")

    upsell_r = sum_row + 6
    ws.merge_cells(f"A{upsell_r}:D{upsell_r}")
    ws.row_dimensions[upsell_r].height = 36
    s(upsell_r, 1,
      "⭐  Want the FULL version? Visit nasritools.etsy.com — 10 templates from €4.99",
      bold=True, bg=light_green, fg=green, size=12, align="center")

    path = OUT_DIR / "NasriTools_Budget_Tracker_FREE.xlsx"
    wb.save(path)
    return path


def create_etsy_listing(token, thumb_img, file_path):
    listing_data = {
        "quantity":   999,
        "title":      "FREE Budget Tracker Spreadsheet Google Sheets | Monthly Budget Planner | Expense Tracker | Free Template | Instant Download",
        "description": """🎁 100% FREE — No email, no signup required. Just download and start.

This free Budget Tracker helps you track income, expenses, and savings every month — automatically.

━━━━━━━━━━━━━━━━━━━━━━━━━━
💰 WHAT'S IN THE FREE VERSION
━━━━━━━━━━━━━━━━━━━━━━━━━━
✔ Monthly Budget Sheet
✔ Income Tracker (3 sources)
✔ Expense Tracker (8 categories)
✔ Auto-Calculating Net Savings
✔ Works on Google Sheets (free) and Excel

━━━━━━━━━━━━━━━━━━━━━━━━━━
⭐ WANT THE FULL VERSION?
━━━━━━━━━━━━━━━━━━━━━━━━━━
The full Budget Tracker includes:
✔ 12-month tracking (full year, one file)
✔ 50+ expense categories
✔ Savings Goal Tracker
✔ Bill Payment Tracker
✔ Net Worth Calculator
✔ Annual Overview Dashboard

👉 Search "NasriTools" on Etsy for the full version!

━━━━━━━━━━━━━━━━━━━━━━━━━━
📥 HOW TO USE
━━━━━━━━━━━━━━━━━━━━━━━━━━
1. Download this file after purchase (€0)
2. Open in Google Sheets
3. File → Make a Copy
4. Enter your numbers and start tracking!

No technical skills needed. Setup: Under 2 minutes.

Questions? Message us on Etsy ♥
nasritools.etsy.com""",
        "price":       0.20,
        "who_made":    "i_did",
        "when_made":   "made_to_order",
        "taxonomy_id": 2078,
        "type":        "download",
        "tags":        [
            "free budget tracker", "free template", "budget spreadsheet",
            "google sheets free", "expense tracker free", "free planner",
            "budget tracker", "monthly budget", "free download",
            "personal finance free", "free spreadsheet", "budget free",
            "free google sheets"
        ],
        "is_digital":  True,
        "file_data":   "",
    }

    r = requests.post(
        f"https://api.etsy.com/v3/application/shops/{SHOP_ID}/listings",
        headers={**auth_headers(token), "Content-Type": "application/json"},
        json=listing_data,
        timeout=30,
    )
    if not r.ok:
        print(f"  [error] create listing: {r.status_code}: {r.text[:200]}")
        return None

    lid = r.json()["listing_id"]
    print(f"  Listing created: {lid}")
    time.sleep(2)

    # Upload image
    buf = io.BytesIO()
    thumb_img.save(buf, "JPEG", quality=93)
    buf.seek(0)
    ri = requests.post(
        f"https://api.etsy.com/v3/application/shops/{SHOP_ID}/listings/{lid}/images",
        headers=auth_headers(token),
        files={"image": ("free_thumbnail.jpg", buf, "image/jpeg")},
        data={"rank": 1},
        timeout=60,
    )
    print(f"  Image upload: {'✓' if ri.ok else '✗ ' + str(ri.status_code)}")
    time.sleep(3)

    # Upload xlsx
    with open(file_path, "rb") as f:
        rf = requests.post(
            f"https://api.etsy.com/v3/application/shops/{SHOP_ID}/listings/{lid}/files",
            headers=auth_headers(token),
            files={"file": (file_path.name, f,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"name": file_path.name, "rank": 1},
            timeout=60,
        )
    print(f"  File upload:  {'✓' if rf.ok else '✗ ' + str(rf.status_code)}")
    time.sleep(3)

    # Publish
    rp = requests.patch(
        f"https://api.etsy.com/v3/application/shops/{SHOP_ID}/listings/{lid}",
        headers={**auth_headers(token), "Content-Type": "application/json"},
        json={"state": "active"},
        timeout=30,
    )
    print(f"  Publish:      {'✓' if rp.ok else '✗ ' + str(rp.status_code)}")

    return lid


def main():
    token = get_token()

    print(f"\n{'='*60}")
    print(f"  NasriTools - Free Product Creator (Lead Magnet)")
    print(f"{'='*60}\n")

    print("  [1/3] Generating FREE thumbnail…", end=" ")
    thumb = make_free_thumbnail()
    print("✓")

    print("  [2/3] Generating FREE Budget Tracker xlsx…", end=" ")
    xlsx_path = make_free_budget_xlsx()
    print(f"✓  ({xlsx_path.stat().st_size // 1024}KB)")

    print("  [3/3] Creating Etsy listing…")
    lid = create_etsy_listing(token, thumb, xlsx_path)

    if lid:
        print(f"\n  ✓ Free product LIVE → https://www.etsy.com/listing/{lid}")
    else:
        print("\n  ✗ Listing creation failed")

    print(f"\n{'='*60}")
    print(f"  Done!")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
