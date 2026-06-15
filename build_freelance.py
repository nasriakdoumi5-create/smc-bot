"""
FREELANCE BUSINESS MANAGER — Complete Google Sheets System
Product 2 for NasriTools Etsy Store
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
import os

# ── Colors ───────────────────────────────────────────
C_DARK    = "0D1B2A"
C_NAV     = "1B2A4A"
C_ACCENT  = "4361EE"
C_GREEN   = "06D6A0"
C_RED     = "EF233C"
C_YELLOW  = "FFD166"
C_ORANGE  = "FB8500"
C_WHITE   = "FFFFFF"
C_LIGHT   = "F8F9FA"
C_GRAY    = "ADB5BD"
C_PURPLE  = "7209B7"

def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False, size=11, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")
def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)
def medium_border():
    s = Side(style="medium", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, val, bg=C_NAV, fg=C_WHITE, sz=11, bold=True):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill(bg); c.font = font(bold=bold, size=sz, color=fg)
    c.alignment = align(); c.border = thin_border()
    return c

def cel(ws, row, col, val=None, bg=C_WHITE, fg="222222",
        bold=False, h="center", wrap=False, sz=11):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill(bg); c.font = font(bold=bold, size=sz, color=fg)
    c.alignment = align(h=h, wrap=wrap); c.border = thin_border()
    return c

def inp(ws, row, col, val=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill(C_YELLOW); c.font = font(size=11)
    c.alignment = align(); c.border = medium_border()
    return c

def set_col(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def set_row(ws, row, h):
    ws.row_dimensions[row].height = h

def banner(ws, r1, c1, r2, c2, text, bg=C_DARK, fg=C_WHITE, sz=16):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=text)
    c.fill = fill(bg); c.font = font(bold=True, size=sz, color=fg)
    c.alignment = align()


# ══════════════════════════════════════════════════════
# TAB 1 — DASHBOARD
# ══════════════════════════════════════════════════════
def build_dashboard(wb):
    ws = wb.active
    ws.title = "🏠 Dashboard"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FF" + C_ACCENT

    for c,w in enumerate([2,22,18,18,18,18,18,2],1):
        set_col(ws, c, w)

    # Title banner
    set_row(ws, 1, 10); set_row(ws, 2, 55); set_row(ws, 3, 14)
    ws.merge_cells("B2:G2")
    t = ws["B2"]
    t.value = "💼  FREELANCE BUSINESS MANAGER  2026"
    t.fill = fill(C_DARK)
    t.font = font(bold=True, size=24, color=C_YELLOW)
    t.alignment = align()

    ws.merge_cells("B3:G3")
    s = ws["B3"]
    s.value = "Your Complete Business Command Center"
    s.fill = fill(C_NAV); s.font = font(size=11, color=C_GRAY, italic=True)
    s.alignment = align()

    # KPI Cards
    kpis = [
        ("💰 MONTHLY REVENUE",  "='💰 Income & Expenses'!C52", C_GREEN,  "B"),
        ("📋 ACTIVE PROJECTS",  "=COUNTIF('📋 Projects'!C4:C53,\"Active\")", C_ACCENT, "D"),
        ("🧾 UNPAID INVOICES",  "=COUNTIF('🧾 Invoices'!G4:G53,\"Unpaid\")", C_RED,    "F"),
    ]

    for title, formula, color, col in kpis:
        nc = ord(col)-64
        set_row(ws, 5, 18); set_row(ws, 6, 42); set_row(ws, 7, 12)
        ws.merge_cells(start_row=5, start_column=nc,
                       end_row=5, end_column=nc+1)
        ws.merge_cells(start_row=6, start_column=nc,
                       end_row=6, end_column=nc+1)
        ws.merge_cells(start_row=7, start_column=nc,
                       end_row=7, end_column=nc+1)

        for r in [5,6,7]:
            c = ws.cell(row=r, column=nc)
            c.fill = fill(color)

        t = ws.cell(row=5, column=nc, value=title)
        t.fill = fill(color); t.font = font(bold=True, size=9, color=C_WHITE)
        t.alignment = align()

        v = ws.cell(row=6, column=nc, value=formula)
        v.fill = fill(color); v.font = font(bold=True, size=28, color=C_WHITE)
        v.alignment = align()
        if "REVENUE" in title: v.number_format = '"$"#,##0'

    # Monthly summary table
    set_row(ws, 9, 26)
    ws.merge_cells("B9:G9")
    banner(ws, 9, 2, 9, 7, "📊  MONTHLY PERFORMANCE OVERVIEW", C_NAV)

    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    hdrs   = ["Month","Revenue","Expenses","Profit","Projects","Invoiced"]
    for i,h in enumerate(hdrs, 2):
        hdr(ws, 10, i, h, C_DARK)
        set_row(ws, 10, 22)

    for mi, month in enumerate(months, 11):
        set_row(ws, mi, 20)
        bg = C_LIGHT if mi%2==0 else C_WHITE
        cel(ws, mi, 2, month, bold=True, bg=bg)
        for c in range(3,8):
            cl = cel(ws, mi, c, 0, bg=bg)
            if c in [3,4,5,7]: cl.number_format = '"$"#,##0'
        # Profit formula
        p = ws.cell(row=mi, column=5, value=f"=C{mi}-D{mi}")
        p.fill = fill(bg); p.font = font(bold=True, color=C_GREEN)
        p.alignment = align(); p.border = thin_border()
        p.number_format = '"$"#,##0'

    # Totals
    set_row(ws, 23, 26)
    cel(ws, 23, 2, "TOTAL", bold=True, bg=C_DARK, fg=C_WHITE)
    for c in range(3,8):
        col_l = get_column_letter(c)
        tc = ws.cell(row=23, column=c,
                     value=f"=SUM({col_l}11:{col_l}22)")
        tc.fill = fill(C_DARK); tc.font = font(bold=True, color=C_YELLOW, size=12)
        tc.alignment = align(); tc.border = thin_border()
        if c != 6: tc.number_format = '"$"#,##0'

    # Top clients section
    set_row(ws, 25, 26)
    ws.merge_cells("B25:G25")
    banner(ws, 25, 2, 25, 7, "⭐  TOP CLIENTS THIS MONTH", C_PURPLE)

    top_hdrs = ["Client Name","Projects","Revenue","Status","Last Contact"]
    for i,h in enumerate(top_hdrs, 2):
        hdr(ws, 26, i, h)
        set_row(ws, 26, 22)

    for r in range(27, 33):
        set_row(ws, r, 22)
        bg = C_LIGHT if r%2==0 else C_WHITE
        for c in range(2,7):
            if c in [3,4,5,6]: inp(ws, r, c) if c!=4 else cel(ws,r,c,bg=bg)
            else: inp(ws, r, c)

    # Tips
    set_row(ws, 34, 20)
    ws.merge_cells("B34:G34")
    tip = ws["B34"]
    tip.value = "💡  Yellow cells = Enter your data  |  Green cells = Auto-calculated  |  Update monthly for accurate reports"
    tip.fill = fill("FFF9C4"); tip.font = font(size=10, color="5D4037", italic=True)
    tip.alignment = align(h="left")


# ══════════════════════════════════════════════════════
# TAB 2 — CLIENT DATABASE
# ══════════════════════════════════════════════════════
def build_clients(wb):
    ws = wb.create_sheet("👥 Clients")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FF" + C_PURPLE

    widths = [2,25,20,20,18,15,15,18,15,18,2]
    for i,w in enumerate(widths,1): set_col(ws,i,w)

    set_row(ws, 1, 48)
    ws.merge_cells("B1:J1")
    banner(ws,1,2,1,10,"👥  CLIENT DATABASE",C_PURPLE,sz=20)

    hdrs = ["Client Name","Company","Email","Phone","Country",
            "Status","Total Projects","Total Revenue","Last Contact"]
    for i,h in enumerate(hdrs,2): hdr(ws,3,i,h); set_row(ws,3,24)

    statuses = ["Active","Inactive","Lead","VIP","Blocked"]
    dv = DataValidation(type="list",
                        formula1='"Active,Inactive,Lead,VIP,Blocked"',
                        allow_blank=True)
    ws.add_data_validation(dv)

    samples = [
        ("TechStart Inc","David Miller","david@techstart.com","+1-555-0101","USA","VIP"),
        ("Creative Agency","Sarah Johnson","sarah@creative.co","+44-20-1234","UK","Active"),
        ("EcomStore",   "Carlos Rivera","carlos@ecom.es",    "+34-91-1234","Spain","Active"),
        ("FoodBrand",   "Amira Hassan","amira@food.ae",     "+971-55-1234","UAE","Lead"),
        ("Startup Hub", "Liu Wei",     "liu@startup.cn",    "+86-10-1234","China","Active"),
    ]

    for ri,(name,company,email,phone,country,status) in enumerate(samples,4):
        set_row(ws,ri,24)
        bg = C_LIGHT if ri%2==0 else C_WHITE
        cel(ws,ri,2,name,bold=True,bg=bg,h="left")
        cel(ws,ri,3,company,bg=bg,h="left")
        inp(ws,ri,4,email)
        inp(ws,ri,5,phone)
        inp(ws,ri,6,country)
        sc = inp(ws,ri,7,status)
        dv.add(sc)
        # Total projects formula
        tp = ws.cell(row=ri,column=8,
            value=f'=COUNTIF(\'📋 Projects\'!D4:D53,B{ri})')
        tp.fill=fill(bg); tp.font=font(color=C_ACCENT,bold=True)
        tp.alignment=align(); tp.border=thin_border()
        # Total revenue formula
        tr = ws.cell(row=ri,column=9,
            value=f'=SUMIF(\'📋 Projects\'!D4:D53,B{ri},\'📋 Projects\'!H4:H53)')
        tr.fill=fill(bg); tr.font=font(color=C_GREEN,bold=True)
        tr.alignment=align(); tr.border=thin_border()
        tr.number_format='"$"#,##0'
        inp(ws,ri,10)

    # Empty rows
    for r in range(9,54):
        set_row(ws,r,22)
        bg=C_LIGHT if r%2==0 else C_WHITE
        for c in range(2,10): cel(ws,r,c,bg=bg)
        for c in [4,5,6,7,10]: inp(ws,r,c)
        ws.cell(row=r,column=8,
            value=f'=IF(B{r}="","",COUNTIF(\'📋 Projects\'!D4:D53,B{r}))')
        ws.cell(row=r,column=9,
            value=f'=IF(B{r}="","",SUMIF(\'📋 Projects\'!D4:D53,B{r},\'📋 Projects\'!H4:H53))')


# ══════════════════════════════════════════════════════
# TAB 3 — PROJECT TRACKER
# ══════════════════════════════════════════════════════
def build_projects(wb):
    ws = wb.create_sheet("📋 Projects")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FF" + C_ACCENT

    widths = [2,6,28,22,15,12,12,15,12,14,14,2]
    for i,w in enumerate(widths,1): set_col(ws,i,w)

    set_row(ws,1,48)
    ws.merge_cells("B1:K1")
    banner(ws,1,2,1,11,"📋  PROJECT TRACKER",C_ACCENT,sz=20)

    hdrs = ["#","Project Name","Client","Category","Status",
            "Start Date","Deadline","Value ($)","% Done","Notes"]
    for i,h in enumerate(hdrs,2): hdr(ws,3,i,h); set_row(ws,3,24)

    statuses = ["Active","Completed","On Hold","Cancelled","Planning"]
    categories = ["Web Design","Development","Marketing","Consulting",
                  "Writing","Video","SEO","Social Media","Other"]

    dv_s = DataValidation(type="list",
                          formula1='"Active,Completed,On Hold,Cancelled,Planning"',
                          allow_blank=True)
    dv_c = DataValidation(type="list",
                          formula1='"Web Design,Development,Marketing,Consulting,Writing,Video,SEO,Social Media,Other"',
                          allow_blank=True)
    ws.add_data_validation(dv_s)
    ws.add_data_validation(dv_c)

    sample_projects = [
        ("Website Redesign","TechStart Inc","Web Design","Active","2026-01-15","2026-03-01",3500,75),
        ("SEO Campaign","Creative Agency","SEO","Active","2026-02-01","2026-04-30",1200,40),
        ("App Development","EcomStore","Development","Active","2026-01-20","2026-05-15",8000,25),
        ("Content Writing","FoodBrand","Writing","Completed","2025-12-01","2026-01-31",600,100),
        ("Social Media","Startup Hub","Social Media","Planning","2026-03-01","2026-06-30",900,0),
    ]

    for ri,(name,client,cat,status,start,end,val,pct) in enumerate(sample_projects,4):
        set_row(ws,ri,24)
        bg=C_LIGHT if ri%2==0 else C_WHITE
        cel(ws,ri,2,ri-3,bg=bg)
        cel(ws,ri,3,name,bold=True,bg=bg,h="left")
        inp(ws,ri,4,client)
        sc=inp(ws,ri,5,cat); dv_c.add(sc)
        ss=inp(ws,ri,6,status); dv_s.add(ss)
        inp(ws,ri,7,start)
        inp(ws,ri,8,end)
        vc=inp(ws,ri,9,val); vc.number_format='"$"#,##0'
        pc=inp(ws,ri,10,pct/100); pc.number_format='0%'
        cel(ws,ri,11,bg=bg,wrap=True,h="left")

        # Color status
        if status=="Completed":
            ws.cell(row=ri,column=6).fill=fill("D4EDDA")
            ws.cell(row=ri,column=6).font=font(bold=True,color="155724")
        elif status=="Active":
            ws.cell(row=ri,column=6).fill=fill("CCE5FF")
            ws.cell(row=ri,column=6).font=font(bold=True,color="004085")

    for r in range(9,54):
        set_row(ws,r,22)
        bg=C_LIGHT if r%2==0 else C_WHITE
        cel(ws,r,2,r-3,bg=bg)
        for c in range(3,12): cel(ws,r,c,bg=bg)
        for c in [4,5,6,7,8,10,11]: inp(ws,r,c)
        iv=inp(ws,r,9,0); iv.number_format='"$"#,##0'

    # Summary row
    set_row(ws,54,28)
    cel(ws,54,2,"",bg=C_DARK)
    ws.merge_cells("C54:H54")
    cel(ws,54,3,"TOTALS",bold=True,bg=C_DARK,fg=C_WHITE)
    tv=ws.cell(row=54,column=9,value="=SUM(I4:I53)")
    tv.fill=fill(C_DARK); tv.font=font(bold=True,color=C_YELLOW,size=14)
    tv.alignment=align(); tv.number_format='"$"#,##0'
    ap=ws.cell(row=54,column=10,value="=AVERAGE(J4:J53)")
    ap.fill=fill(C_DARK); ap.font=font(bold=True,color=C_GREEN)
    ap.alignment=align(); ap.number_format='0%'


# ══════════════════════════════════════════════════════
# TAB 4 — INVOICE GENERATOR
# ══════════════════════════════════════════════════════
def build_invoices(wb):
    ws = wb.create_sheet("🧾 Invoices")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FF" + C_ORANGE

    widths = [2,12,22,22,18,14,14,14,16,2]
    for i,w in enumerate(widths,1): set_col(ws,i,w)

    set_row(ws,1,48)
    ws.merge_cells("B1:I1")
    banner(ws,1,2,1,9,"🧾  INVOICE TRACKER",C_ORANGE,sz=20)

    hdrs = ["Invoice #","Client","Project","Amount","Date Sent","Due Date","Status","Days Overdue"]
    for i,h in enumerate(hdrs,2): hdr(ws,3,i,h); set_row(ws,3,24)

    statuses = ["Paid","Unpaid","Overdue","Partial","Cancelled"]
    dv = DataValidation(type="list",
                        formula1='"Paid,Unpaid,Overdue,Partial,Cancelled"',
                        allow_blank=True)
    ws.add_data_validation(dv)

    samples = [
        ("INV-001","TechStart Inc","Website Redesign",1750,"2026-01-15","2026-01-29","Paid"),
        ("INV-002","Creative Agency","SEO Campaign",600,"2026-02-01","2026-02-15","Unpaid"),
        ("INV-003","EcomStore","App Development",2000,"2026-02-10","2026-02-24","Overdue"),
        ("INV-004","FoodBrand","Content Writing",600,"2026-01-31","2026-02-14","Paid"),
        ("INV-005","Startup Hub","Social Media",450,"2026-03-01","2026-03-15","Unpaid"),
    ]

    for ri,(inv,client,proj,amt,sent,due,status) in enumerate(samples,4):
        set_row(ws,ri,24)
        bg=C_LIGHT if ri%2==0 else C_WHITE
        cel(ws,ri,2,inv,bold=True,bg=bg)
        inp(ws,ri,3,client)
        inp(ws,ri,4,proj)
        ac=inp(ws,ri,5,amt); ac.number_format='"$"#,##0.00'
        inp(ws,ri,6,sent)
        inp(ws,ri,7,due)
        sc=inp(ws,ri,8,status); dv.add(sc)
        # Days overdue
        od=ws.cell(row=ri,column=9,
            value=f'=IF(H{ri}="Paid","✅",IF(TODAY()>G{ri},TODAY()-G{ri},"On time"))')
        od.fill=fill(bg); od.font=font(bold=True)
        od.alignment=align(); od.border=thin_border()

        # Color status
        colors = {"Paid":("D4EDDA","155724"),"Unpaid":("FFF3CD","856404"),
                  "Overdue":("F8D7DA","721C24"),"Partial":("D1ECF1","0C5460")}
        if status in colors:
            bg2,fg2 = colors[status]
            ws.cell(row=ri,column=8).fill=fill(bg2)
            ws.cell(row=ri,column=8).font=font(bold=True,color=fg2)

    for r in range(9,54):
        set_row(ws,r,22)
        bg=C_LIGHT if r%2==0 else C_WHITE
        for c in range(2,9): cel(ws,r,c,bg=bg)
        for c in [3,4,6,7]: inp(ws,r,c)
        av=inp(ws,r,5,0); av.number_format='"$"#,##0.00'
        sv=inp(ws,r,8,"Unpaid"); dv.add(sv)
        ws.cell(row=r,column=9,
            value=f'=IF(E{r}=0,"",IF(H{r}="Paid","✅",IF(TODAY()>G{r},TODAY()-G{r},"On time")))')

    # Summary
    set_row(ws,55,26)
    ws.merge_cells("B55:D55")
    cel(ws,55,2,"SUMMARY",bold=True,bg=C_DARK,fg=C_WHITE)
    summaries = [
        (5,"Total Invoiced","=SUM(E4:E53)"),
        (6,"Total Paid",'=SUMIF(H4:H53,"Paid",E4:E53)'),
        (7,"Total Unpaid",'=SUMIF(H4:H53,"Unpaid",E4:E53)'),
        (8,"Total Overdue",'=SUMIF(H4:H53,"Overdue",E4:E53)'),
    ]
    for col,label,formula in summaries:
        cel(ws,55,col,label,bold=True,bg=C_NAV,fg=C_WHITE)
        vc=ws.cell(row=56,column=col,value=formula)
        vc.fill=fill(C_DARK); vc.font=font(bold=True,color=C_YELLOW,size=13)
        vc.alignment=align(); vc.number_format='"$"#,##0.00'
        set_row(ws,56,30)


# ══════════════════════════════════════════════════════
# TAB 5 — INCOME & EXPENSES
# ══════════════════════════════════════════════════════
def build_income(wb):
    ws = wb.create_sheet("💰 Income & Expenses")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FF" + C_GREEN

    widths = [2,12,22,22,18,16,16,2]
    for i,w in enumerate(widths,1): set_col(ws,i,w)

    set_row(ws,1,48)
    ws.merge_cells("B1:G1")
    banner(ws,1,2,1,7,"💰  INCOME & EXPENSE TRACKER",C_GREEN[:-2]+"88",sz=20)

    # C52 referenced by Dashboard
    ws["C52"] = "=SUMIF(F4:F51,\"Income\",E4:E51)"
    ws["C52"].font = font(size=1,color=C_WHITE)

    hdrs = ["Date","Description","Client/Vendor","Amount","Type","Category"]
    for i,h in enumerate(hdrs,2): hdr(ws,3,i,h); set_row(ws,3,24)

    types = ["Income","Expense"]
    cats  = ["Project Payment","Retainer","Bonus","Software","Hardware",
             "Marketing","Office","Travel","Tax","Other"]

    dv_t = DataValidation(type="list",formula1='"Income,Expense"',allow_blank=True)
    dv_c = DataValidation(type="list",
                          formula1='"Project Payment,Retainer,Bonus,Software,Hardware,Marketing,Office,Travel,Tax,Other"',
                          allow_blank=True)
    ws.add_data_validation(dv_t)
    ws.add_data_validation(dv_c)

    samples = [
        ("2026-01-29","Website Redesign P1","TechStart Inc",1750,"Income","Project Payment"),
        ("2026-01-31","Adobe CC","Adobe",54,"Expense","Software"),
        ("2026-02-14","Content Writing","FoodBrand",600,"Income","Project Payment"),
        ("2026-02-15","Office Supplies","Amazon",35,"Expense","Office"),
        ("2026-02-20","SEO Tools","Ahrefs",99,"Expense","Software"),
        ("2026-03-01","Retainer Fee","Creative Agency",500,"Income","Retainer"),
    ]

    for ri,(date,desc,cv,amt,typ,cat) in enumerate(samples,4):
        set_row(ws,ri,22)
        bg=C_LIGHT if ri%2==0 else C_WHITE
        inp(ws,ri,2,date)
        inp(ws,ri,3,desc)
        inp(ws,ri,4,cv)
        ac=inp(ws,ri,5,amt); ac.number_format='"$"#,##0.00'
        tc=inp(ws,ri,6,typ); dv_t.add(tc)
        cc=inp(ws,ri,7,cat); dv_c.add(cc)

        clr = C_GREEN if typ=="Income" else C_RED
        ws.cell(row=ri,column=5).number_format='"$"#,##0.00'
        ws.cell(row=ri,column=6).fill=fill("D4EDDA" if typ=="Income" else "F8D7DA")
        ws.cell(row=ri,column=6).font=font(bold=True,
            color="155724" if typ=="Income" else "721C24")

    for r in range(10,52):
        set_row(ws,r,22)
        bg=C_LIGHT if r%2==0 else C_WHITE
        for c in range(2,8): cel(ws,r,c,bg=bg)
        for c in [2,3,4]: inp(ws,r,c)
        av=inp(ws,r,5,0); av.number_format='"$"#,##0.00'
        tv=inp(ws,r,6,"Income"); dv_t.add(tv)
        cv=inp(ws,r,7,"Project Payment"); dv_c.add(cv)

    # Summary
    set_row(ws,53,28)
    ws.merge_cells("B53:C53")
    cel(ws,53,2,"TOTALS",bold=True,bg=C_DARK,fg=C_WHITE)
    tots = [
        (4,"Total Income",'=SUMIF(F4:F51,"Income",E4:E51)',C_GREEN),
        (5,"Total Expenses",'=SUMIF(F4:F51,"Expense",E4:E51)',C_RED),
        (6,"Net Profit","=D54-E54",C_ACCENT),
        (7,"Tax Est. (25%)",'=SUMIF(F4:F51,"Income",E4:E51)*0.25',C_ORANGE),
    ]
    for col,label,formula,color in tots:
        cel(ws,53,col,label,bold=True,bg=C_NAV,fg=C_WHITE)
        vc=ws.cell(row=54,column=col,value=formula)
        vc.fill=fill(C_DARK); vc.font=font(bold=True,color=color,size=13)
        vc.alignment=align(); vc.number_format='"$"#,##0.00'
        set_row(ws,54,30)


# ══════════════════════════════════════════════════════
# TAB 6 — TIME TRACKER
# ══════════════════════════════════════════════════════
def build_time(wb):
    ws = wb.create_sheet("⏱️ Time Tracker")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FF" + C_YELLOW

    widths = [2,12,28,22,10,10,12,16,18,2]
    for i,w in enumerate(widths,1): set_col(ws,i,w)

    set_row(ws,1,48)
    ws.merge_cells("B1:I1")
    banner(ws,1,2,1,9,"⏱️  TIME TRACKER",C_NAV,sz=20)

    hdrs = ["Date","Task","Client","Start","End","Hours","Rate ($/hr)","Earned"]
    for i,h in enumerate(hdrs,2): hdr(ws,3,i,h); set_row(ws,3,24)

    samples = [
        ("2026-02-01","Homepage Design","TechStart Inc","09:00","13:00",4,85),
        ("2026-02-01","Client Meeting","Creative Agency","14:00","15:00",1,85),
        ("2026-02-02","SEO Audit","Creative Agency","09:00","12:30",3.5,75),
        ("2026-02-03","App Wireframes","EcomStore","10:00","18:00",8,95),
        ("2026-02-04","Content Writing","FoodBrand","09:00","12:00",3,65),
    ]

    for ri,(date,task,client,start,end,hrs,rate) in enumerate(samples,4):
        set_row(ws,ri,22)
        bg=C_LIGHT if ri%2==0 else C_WHITE
        inp(ws,ri,2,date)
        inp(ws,ri,3,task)
        inp(ws,ri,4,client)
        inp(ws,ri,5,start)
        inp(ws,ri,6,end)
        # Hours formula
        hc=ws.cell(row=ri,column=7,
            value=f"=IF(E{ri}=\"\",0,(TIMEVALUE(F{ri})-TIMEVALUE(E{ri}))*24)")
        hc.fill=fill(bg); hc.font=font(bold=True,color=C_ACCENT)
        hc.alignment=align(); hc.border=thin_border()
        hc.number_format='0.0'
        rc=inp(ws,ri,8,rate); rc.number_format='"$"#,##0'
        # Earned formula
        ec=ws.cell(row=ri,column=9,value=f"=G{ri}*H{ri}")
        ec.fill=fill(bg); ec.font=font(bold=True,color=C_GREEN)
        ec.alignment=align(); ec.border=thin_border()
        ec.number_format='"$"#,##0.00'

    for r in range(9,52):
        set_row(ws,r,22)
        bg=C_LIGHT if r%2==0 else C_WHITE
        for c in range(2,9): cel(ws,r,c,bg=bg)
        for c in [2,3,4,5,6]: inp(ws,r,c)
        rv=inp(ws,r,8,0); rv.number_format='"$"#,##0'
        ws.cell(row=r,column=7,
            value=f'=IF(E{r}="",0,(TIMEVALUE(F{r})-TIMEVALUE(E{r}))*24)')
        ws.cell(row=r,column=9,value=f"=G{r}*H{r}")
        ws.cell(row=r,column=9).number_format='"$"#,##0.00'

    # Summary
    set_row(ws,53,28)
    ws.merge_cells("B53:F53")
    cel(ws,53,2,"TOTALS",bold=True,bg=C_DARK,fg=C_WHITE)
    th=ws.cell(row=53,column=7,value="=SUM(G4:G51)")
    th.fill=fill(C_DARK); th.font=font(bold=True,color=C_YELLOW,size=13)
    th.alignment=align(); th.number_format='0.0 "hrs"'
    te=ws.cell(row=53,column=9,value="=SUM(I4:I51)")
    te.fill=fill(C_DARK); te.font=font(bold=True,color=C_GREEN,size=13)
    te.alignment=align(); te.number_format='"$"#,##0.00'


# ══════════════════════════════════════════════════════
# TAB 7 — REPORTS
# ══════════════════════════════════════════════════════
def build_reports(wb):
    ws = wb.create_sheet("📊 Reports")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FF" + C_RED

    widths = [2,28,20,20,2]
    for i,w in enumerate(widths,1): set_col(ws,i,w)

    set_row(ws,1,48)
    ws.merge_cells("B1:D1")
    banner(ws,1,2,1,4,"📊  AUTO REPORTS",C_RED,sz=20)

    sections = [
        ("💰 FINANCIAL SUMMARY",C_GREEN,"B3"),
        ("📋 PROJECT SUMMARY",  C_ACCENT,"B14"),
        ("🧾 INVOICE SUMMARY",  C_ORANGE,"B25"),
        ("⏱️ TIME SUMMARY",     C_PURPLE,"B36"),
    ]

    metrics = [
        [("Total Revenue","='💰 Income & Expenses'!D54"),
         ("Total Expenses","='💰 Income & Expenses'!E54"),
         ("Net Profit","='💰 Income & Expenses'!F54"),
         ("Tax Estimate","='💰 Income & Expenses'!G54"),
         ("Profit Margin %","=IF(D4=0,0,F4/D4*100)"),
         ("Monthly Average","=F4/12"),
         ("Best Month","=MAX('🏠 Dashboard'!C11:C22)"),
         ("Worst Month","=MIN(IF('🏠 Dashboard'!C11:C22>0,'🏠 Dashboard'!C11:C22))"),],
        [("Total Projects","=COUNTA('📋 Projects'!C4:C53)"),
         ("Active Projects","=COUNTIF('📋 Projects'!F4:F53,\"Active\")"),
         ("Completed","=COUNTIF('📋 Projects'!F4:F53,\"Completed\")"),
         ("Total Project Value","=SUM('📋 Projects'!I4:I53)"),
         ("Avg Project Value","=IFERROR(D28/D25,0)"),
         ("Completion Rate %","=IFERROR(D26/D25*100,0)"),],
        [("Total Invoiced","=SUM('🧾 Invoices'!E4:E53)"),
         ("Total Paid","=SUMIF('🧾 Invoices'!H4:H53,\"Paid\",'🧾 Invoices'!E4:E53)"),
         ("Unpaid","=SUMIF('🧾 Invoices'!H4:H53,\"Unpaid\",'🧾 Invoices'!E4:E53)"),
         ("Overdue","=SUMIF('🧾 Invoices'!H4:H53,\"Overdue\",'🧾 Invoices'!E4:E53)"),
         ("Collection Rate %","=IFERROR(D38/D37*100,0)"),],
        [("Total Hours Tracked","=SUM('⏱️ Time Tracker'!G4:G51)"),
         ("Total Earned","=SUM('⏱️ Time Tracker'!I4:I51)"),
         ("Avg Hourly Rate","=IFERROR(D49/D48,0)"),
         ("Most Productive Day","=TEXT(INDEX('⏱️ Time Tracker'!B4:B51,MATCH(MAX('⏱️ Time Tracker'!G4:G51),'⏱️ Time Tracker'!G4:G51,0)),\"DD/MM/YYYY\")"),],
    ]

    row = 3
    for (sec_title, color, _), met_list in zip(sections, metrics):
        set_row(ws, row, 28)
        ws.merge_cells(f"B{row}:D{row}")
        banner(ws, row, 2, row, 4, sec_title, color, sz=14)
        row += 1

        for label, formula in met_list:
            set_row(ws, row, 24)
            bg = C_LIGHT if row%2==0 else C_WHITE
            cel(ws, row, 2, label, bold=True, bg=bg, h="left")
            vc = ws.cell(row=row, column=3, value=formula)
            vc.fill = fill(bg)
            vc.font = font(bold=True, color=color, size=12)
            vc.alignment = align()
            vc.border = thin_border()
            if "$" in label or "Revenue" in label or "Earned" in label or \
               "Profit" in label or "Tax" in label or "Value" in label or \
               "Paid" in label or "Invoiced" in label or "Unpaid" in label or \
               "Overdue" in label or "Rate" in label and "%" not in label:
                vc.number_format = '"$"#,##0.00'
            elif "%" in label:
                vc.number_format = '0.00"%"'
            elif "Hours" in label:
                vc.number_format = '0.0 "hrs"'
            cel(ws, row, 4, bg=bg)
            row += 1
        row += 1


# ══════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    build_dashboard(wb)
    build_clients(wb)
    build_projects(wb)
    build_invoices(wb)
    build_income(wb)
    build_time(wb)
    build_reports(wb)

    out = "/home/user/smc-bot/Freelance_Business_Manager.xlsx"
    wb.save(out)
    import os
    print(f"✅ Created: {out}")
    print(f"📦 Size: {os.path.getsize(out)//1024} KB")
    print("📋 Tabs: Dashboard | Clients | Projects | Invoices | Income & Expenses | Time Tracker | Reports")

if __name__ == "__main__":
    main()
