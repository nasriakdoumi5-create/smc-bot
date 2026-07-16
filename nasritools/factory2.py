"""
NasriTools Factory v2 — quality workbook generator.
Reuses the existing CATALOG configs but renders REAL templates:
  • START HERE tab with styled instructions per product
  • Dashboard with KPI cards + "what's inside" guide
  • Data tables: 12+ varied realistic sample rows, dropdowns,
    banded rows, borders, currency formats
  • Reports tabs with styled summary sections
All formulas Google Sheets compatible.
"""
import re
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import DataBarRule

OUTPUT_DIR = Path(__file__).parent.parent / "output"

WHITE = "FFFFFF"
CREAM = "FFF7F0"
GREY_B = "DDDDDD"

THIN   = Side(style="thin", color=GREY_B)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP   = Alignment(horizontal="left", vertical="center", wrap_text=True)

F_TD   = Font(name="Arial", size=10, color="222222")
F_NOTE = Font(name="Arial", italic=True, size=10, color="777777")

MONEY_WORDS = ("amount", "price", "cost", "budget", "total", "revenue",
               "income", "spent", "paid", "value", "fee", "salary",
               "profit", "expense", "balance", "target", "saved", "rate/hr")
PCT_WORDS   = ("progress", "% ", "percent", "rate")

# vocab banks for varied sample rows
VOCAB = {
    "person":  ["Alex Martin", "Sara Lopez", "Omar Haddad", "Emma Fischer",
                "Liam Carter", "Nadia Rahali", "Tom Becker", "Julia Romero",
                "Karim Ziani", "Anna Kowalska", "David Chen", "Lea Moreau"],
    "company": ["BrightWave Studio", "GreenLeaf Co.", "Nova Digital",
                "Atlas Consulting", "Pixel & Co", "Blue Harbor Ltd",
                "Sunrise Media", "UrbanCraft", "PrimeWorks", "EchoLab",
                "Vertex Group", "ClearPath Inc"],
    "task":    ["Kickoff meeting", "Send proposal", "Review feedback",
                "Update website copy", "Prepare invoice", "Client call",
                "Design first draft", "Fix reported bug", "Weekly report",
                "Plan next sprint", "Publish blog post", "Team sync"],
    "note":    ["", "Paid", "Urgent", "Follow up Friday", "", "Monthly",
                "Confirmed", "", "Waiting reply", "Done early", "", "Recheck"],
}

def _vocab_for(colname: str):
    n = colname.lower()
    if any(w in n for w in ("client", "name", "student", "employee", "member",
                            "contact", "guest", "customer", "lead", "author")):
        return VOCAB["person"]
    if any(w in n for w in ("company", "vendor", "supplier", "brand", "sponsor")):
        return VOCAB["company"]
    if any(w in n for w in ("task", "item", "activity", "action", "step",
                            "title", "topic", "goal", "milestone", "project")):
        return VOCAB["task"]
    if any(w in n for w in ("note", "comment", "detail", "remark")):
        return VOCAB["note"]
    return None

def _is_money(colname: str) -> bool:
    n = colname.lower()
    return any(w in n for w in MONEY_WORDS)

def _parse_date(s):
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", str(s).strip())
    if not m:
        return None
    y, mo, dd = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return date(y, mo, dd)
    except ValueError:
        return None

def _num(s):
    try:
        return float(str(s).replace(",", "").replace("€", "").replace("$", ""))
    except (ValueError, TypeError):
        return None

def _hex(cfg, key, default):
    v = str(cfg.get(key, default) or default).lstrip("#")
    return v if re.match(r"^[0-9A-Fa-f]{6}$", v) else default

def _sample_value(col, row_i, base_date):
    """Deterministic varied sample for row row_i (0-based)."""
    name = col.get("name", "")
    if col.get("dropdown"):
        opts = col["dropdown"]
        return opts[row_i % len(opts)]
    sample = col.get("sample", "")
    d = _parse_date(sample)
    if d is not None:
        return (base_date + timedelta(days=(row_i * 5 + (row_i % 3)) % 28))
    n = _num(sample)
    if n is not None:
        factors = [1.0, 0.7, 1.3, 0.85, 1.15, 0.6, 1.45, 0.95, 1.25, 0.8, 1.1, 0.75]
        val = n * factors[row_i % len(factors)]
        return round(val, 2) if n < 100 else round(val)
    # numbered IDs like "INV-001", "#004", "W1" → increment per row
    m = re.match(r"^(.*?)(\d+)$", str(sample).strip())
    if m and m.group(2):
        width = len(m.group(2))
        return f"{m.group(1)}{int(m.group(2)) + row_i:0{width}d}"
    # emails → derive from person names
    if "@" in str(sample):
        first = VOCAB["person"][row_i % len(VOCAB["person"])].split()[0].lower()
        return f"{first}@mail.com"
    voc = _vocab_for(name)
    if voc:
        return voc[row_i % len(voc)]
    # unknown text column: sample first, then sparse variations
    return sample if row_i % 3 != 2 else ""

def _apply_formula(template, row):
    return template.replace("{ROW}", str(row))


# ═══════════════════════════════════════════════════════════════════════
def build_start_here(wb, cfg, dark, accent):
    ws = wb.create_sheet("START HERE", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col, w in zip("BCDEF", [16, 30, 30, 30, 16]):
        ws.column_dimensions[col].width = w

    fill_dark = PatternFill("solid", fgColor=dark)
    ws.merge_cells("B2:F3")
    c = ws["B2"]
    c.value = f"{cfg.get('listing_emoji','📊')}  {cfg['name'].upper()}"
    c.font = Font(name="Arial", bold=True, size=20, color=WHITE)
    c.fill = fill_dark; c.alignment = CENTER
    ws.merge_cells("B4:F4")
    c = ws["B4"]
    c.value = cfg.get("subtitle", "")
    c.font = Font(name="Arial", size=11, color="DDDDDD")
    c.fill = fill_dark; c.alignment = CENTER

    r = 6
    data_tabs = [t for t in cfg.get("tabs", []) if t.get("type") == "table"]
    for i, t in enumerate(data_tabs, 1):
        ws.merge_cells(f"B{r}:F{r}")
        c = ws[f"B{r}"]
        c.value = f"STEP {i} — Fill in “{t['name']}”"
        c.font = Font(name="Arial", bold=True, size=12, color=accent)
        ws.merge_cells(f"B{r+1}:F{r+1}")
        c = ws[f"B{r+1}"]
        c.value = t.get("description", "")
        c.font = Font(name="Arial", size=11, color="333333")
        c.alignment = WRAP
        ws.row_dimensions[r+1].height = 28
        r += 3

    ws.merge_cells(f"B{r}:F{r}")
    c = ws[f"B{r}"]
    c.value = "Then open the Dashboard — everything calculates automatically."
    c.font = Font(name="Arial", bold=True, size=12, color=accent)
    r += 2

    for tip in [
        "TIP: Delete the sample rows once you've seen how it works.",
        "TIP: On Google Sheets? File → Make a copy first, then edit your copy.",
        "TIP: Columns with dropdowns keep your data clean — use them.",
        "Need help? Message NasriTools on Etsy — I reply fast. 🧡",
    ]:
        ws.merge_cells(f"B{r}:F{r}")
        c = ws[f"B{r}"]; c.value = tip; c.font = F_NOTE
        r += 2


def build_table(wb, tab, dark, n_rows=12):
    ws = wb.create_sheet(tab["name"][:31])
    ws.sheet_view.showGridLines = False
    cols = tab.get("columns", [])
    fill_dark = PatternFill("solid", fgColor=dark)
    fill_band = PatternFill("solid", fgColor=CREAM)

    for i, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=col.get("name", ""))
        c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill = fill_dark; c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = col.get("width", 16)

    base_date = date(2026, 6, 1)
    for r_i in range(n_rows):
        row = 2 + r_i
        for c_i, col in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c_i)
            if col.get("formula"):
                cell.value = _apply_formula(col["formula"], row)
            else:
                v = _sample_value(col, r_i, base_date)
                cell.value = v
                if isinstance(v, date):
                    cell.number_format = "yyyy-mm-dd"
                elif isinstance(v, (int, float)) and _is_money(col.get("name", "")):
                    cell.number_format = '"€"#,##0.00'
            cell.font = F_TD
            cell.border = BORDER
            if row % 2 == 0:
                cell.fill = fill_band

    # dropdown validations for full editing range
    for c_i, col in enumerate(cols, 1):
        if col.get("dropdown"):
            letter = get_column_letter(c_i)
            joined = ",".join(o.replace(",", " ") for o in col["dropdown"])[:250]
            dv = DataValidation(type="list", formula1=f'"{joined}"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{letter}2:{letter}500")


def build_dashboard(wb, cfg, dark, accent, green):
    tab = next((t for t in cfg.get("tabs", []) if t.get("type") == "dashboard"), None)
    ws = wb.create_sheet("Dashboard", 1)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col, w in zip("BCDEFGHI", [15, 15, 15, 15, 15, 15, 15, 15]):
        ws.column_dimensions[col].width = w

    fill_dark = PatternFill("solid", fgColor=dark)
    ws.merge_cells("B2:I2")
    c = ws["B2"]
    c.value = f"📊  {cfg['name'].upper()} — DASHBOARD"
    c.font = Font(name="Arial", bold=True, size=15, color=WHITE)
    c.fill = fill_dark; c.alignment = CENTER
    ws.row_dimensions[2].height = 28

    kpis = (tab or {}).get("kpis", [])[:4]
    card_fills = [accent, green, dark, "6366F1"]
    col_pairs = [("B", "C"), ("D", "E"), ("F", "G"), ("H", "I")]
    for i, kpi in enumerate(kpis):
        c1, c2 = col_pairs[i]
        fill = PatternFill("solid", fgColor=card_fills[i % len(card_fills)])
        ws.merge_cells(f"{c1}4:{c2}4"); ws.merge_cells(f"{c1}5:{c2}5")
        lab = ws[f"{c1}4"]; lab.value = kpi.get("label", "").upper()
        lab.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        lab.fill = fill; lab.alignment = CENTER
        val = ws[f"{c1}5"]; val.value = kpi.get("formula", "")
        val.font = Font(name="Arial", bold=True, size=16, color=WHITE)
        val.fill = fill; val.alignment = CENTER
    ws.row_dimensions[5].height = 26

    # "what's inside" guide
    r = 8
    ws[f"B{r}"] = "WHAT'S INSIDE"
    ws[f"B{r}"].font = Font(name="Arial", bold=True, size=12, color=dark)
    r += 1
    for t in cfg.get("tabs", []):
        if t.get("type") == "dashboard":
            continue
        ws.merge_cells(f"B{r}:I{r}")
        c = ws[f"B{r}"]
        c.value = f"{t.get('emoji','📋')}  {t['name']} — {t.get('description','')}"
        c.font = Font(name="Arial", size=10, color="444444")
        c.alignment = WRAP
        r += 1
    r += 1
    ws.merge_cells(f"B{r}:I{r}")
    c = ws[f"B{r}"]
    c.value = "Numbers update automatically as you fill the other tabs. Start with sample data deleted."
    c.font = F_NOTE


def build_reports(wb, tab, dark, accent):
    ws = wb.create_sheet(tab["name"][:31])
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20

    fill_dark = PatternFill("solid", fgColor=dark)
    ws.merge_cells("B2:C2")
    c = ws["B2"]; c.value = f"📈  {tab['name'].upper()}"
    c.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill = fill_dark; c.alignment = CENTER

    r = 4
    for section in tab.get("sections", []):
        ws.merge_cells(f"B{r}:C{r}")
        c = ws[f"B{r}"]; c.value = section.get("title", "")
        c.font = Font(name="Arial", bold=True, size=12, color=accent)
        r += 1
        for item in section.get("items", []):
            lab = ws.cell(row=r, column=2, value=item.get("label", ""))
            lab.font = F_TD; lab.border = BORDER
            val = ws.cell(row=r, column=3, value=item.get("formula", ""))
            val.font = Font(name="Arial", bold=True, size=11, color="222222")
            val.border = BORDER
            val.number_format = '"€"#,##0.00' if _is_money(item.get("label", "")) else "General"
            if r % 2 == 0:
                lab.fill = PatternFill("solid", fgColor=CREAM)
                val.fill = PatternFill("solid", fgColor=CREAM)
            r += 1
        r += 1


def build_workbook(cfg) -> Path:
    dark   = _hex(cfg, "dark_color", "1A1A2E")
    accent = _hex(cfg, "accent_color", "F97316")
    green  = _hex(cfg, "green_color", "22C55E")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for tab in cfg.get("tabs", []):
        if tab.get("type") == "table":
            build_table(wb, tab, dark)
        elif tab.get("type") == "reports":
            build_reports(wb, tab, dark, accent)

    build_dashboard(wb, cfg, dark, accent, green)
    build_start_here(wb, cfg, dark, accent)

    # order: START HERE, Dashboard, then data tabs
    out = OUTPUT_DIR / cfg["slug"] / f"{cfg['slug']}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
