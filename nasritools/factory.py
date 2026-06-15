"""
NasriTools Product Factory
Given a product config dict, generates:
  1. Excel/Google Sheets template (.xlsx)
  2. 5 Etsy product images (2000x2000 JPG)
  3. PDF guide
  4. Etsy listing text (title, description, tags, price)
All output goes to: output/<slug>/
"""

import os
import json
import textwrap
from pathlib import Path
from datetime import datetime

# ── Pillow ──────────────────────────────────────────────────────────────
from PIL import Image, ImageDraw, ImageFont

# ── openpyxl ────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── reportlab ───────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT

OUTPUT_DIR = Path("/home/user/smc-bot/output")


# ════════════════════════════════════════════════════════════════════════
#  HELPERS
# ════════════════════════════════════════════════════════════════════════

def out_dir(slug: str) -> Path:
    d = OUTPUT_DIR / slug
    d.mkdir(parents=True, exist_ok=True)
    return d


def try_font(size: int, bold: bool = False):
    """Load a PIL font, fall back to default if not available."""
    paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf" if bold else
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSansBold.otf" if bold else
        "/usr/share/fonts/truetype/freefont/FreeSans.otf",
    ]
    for p in paths:
        if os.path.exists(p):
            try:
                return ImageFont.truetype(p, size)
            except Exception:
                pass
    return ImageFont.load_default()


def hex2rgb(h: str):
    h = h.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))


def draw_text_wrapped(draw, text, x, y, max_width, font, fill, line_height=None):
    """Draw word-wrapped text, return y after last line."""
    words = text.split()
    lines = []
    current = ""
    for w in words:
        test = (current + " " + w).strip()
        bbox = draw.textbbox((0, 0), test, font=font)
        if bbox[2] <= max_width:
            current = test
        else:
            if current:
                lines.append(current)
            current = w
    if current:
        lines.append(current)
    lh = line_height or (font.size + 6 if hasattr(font, "size") else 20)
    for line in lines:
        draw.text((x, y), line, font=font, fill=fill)
        y += lh
    return y


# ════════════════════════════════════════════════════════════════════════
#  1. EXCEL BUILDER
# ════════════════════════════════════════════════════════════════════════

# Excel style constants
C = {
    "dark":   "0D1B2A",
    "accent": "4361EE",
    "green":  "06D6A0",
    "red":    "EF233C",
    "yellow": "FFD166",
    "light":  "F8F9FA",
    "white":  "FFFFFF",
    "gray":   "888888",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(hex_color="000000", bold=False, size=11):
    return Font(name="Calibri", color=hex_color, bold=bold, size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, cols, row=1):
    """Write a header row with dark background."""
    for col_idx, (text, width) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.fill = fill(C["dark"])
        cell.font = font(C["white"], bold=True, size=11)
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def input_cell(ws, row, col, value=""):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(C["yellow"])
    cell.font = font(C["dark"])
    cell.alignment = left()
    cell.border = thin_border()
    return cell

def formula_cell(ws, row, col, formula, color=None):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.fill = fill(color or C["light"])
    cell.font = font(C["dark"])
    cell.alignment = center()
    cell.border = thin_border()
    return cell


def build_excel(cfg: dict, slug: str) -> str:
    """Build the Excel template from a product config. Returns file path."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    tabs = cfg.get("tabs", [])
    for tab in tabs:
        ws = wb.create_sheet(tab["name"])
        tab_color = tab.get("color", C["accent"])
        ws.sheet_properties.tabColor = "FF" + tab_color

        tab_type = tab.get("type", "table")

        if tab_type == "dashboard":
            _build_dashboard_tab(ws, tab, cfg)
        elif tab_type == "table":
            _build_table_tab(ws, tab)
        elif tab_type == "reports":
            _build_reports_tab(ws, tab, cfg)
        else:
            _build_table_tab(ws, tab)

    path = str(out_dir(slug) / f"{slug}.xlsx")
    wb.save(path)
    return path


def _build_dashboard_tab(ws, tab, cfg):
    ws.row_dimensions[1].height = 40
    title_cell = ws.cell(row=1, column=1, value=cfg.get("name", "Dashboard"))
    title_cell.fill = fill(C["dark"])
    title_cell.font = font(C["white"], bold=True, size=18)
    title_cell.alignment = center()
    ws.merge_cells("A1:F1")

    kpis = tab.get("kpis", [])
    for i, kpi in enumerate(kpis):
        col = i + 1
        ws.cell(row=3, column=col, value=kpi["label"]).font = font(C["gray"], bold=True, size=9)
        ws.cell(row=3, column=col).alignment = center()
        formula_cell(ws, 4, col, kpi.get("formula", ""), C["accent"])
        ws.cell(row=4, column=col).font = font(C["white"], bold=True, size=14)
        ws.column_dimensions[get_column_letter(col)].width = 20

    ws.row_dimensions[4].height = 36


def _build_table_tab(ws, tab):
    cols = tab.get("columns", [])
    header_row(ws, [(c["name"], c.get("width", 18)) for c in cols])

    sample_rows = tab.get("sample_rows", 5)
    dv_map = {}
    for col_idx, col in enumerate(cols, 1):
        if col.get("dropdown"):
            options = '","'.join(col["dropdown"])
            formula = f'"{options}"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(dv)
            dv_map[col_idx] = dv

    for row_idx in range(2, sample_rows + 2):
        ws.row_dimensions[row_idx].height = 22
        for col_idx, col in enumerate(cols, 1):
            if col.get("formula"):
                f = col["formula"].replace("{ROW}", str(row_idx))
                formula_cell(ws, row_idx, col_idx, f)
            else:
                c = input_cell(ws, row_idx, col_idx, col.get("sample", ""))
                if col_idx in dv_map:
                    dv_map[col_idx].add(c)


def _build_reports_tab(ws, tab, cfg):
    ws.cell(row=1, column=1, value="Reports — Auto-Generated").fill = fill(C["accent"])
    ws.cell(row=1, column=1).font = font(C["white"], bold=True, size=14)
    ws.merge_cells("A1:D1")
    sections = tab.get("sections", [])
    row = 3
    for sec in sections:
        ws.cell(row=row, column=1, value=sec["title"]).font = font(C["accent"], bold=True, size=12)
        row += 1
        for item in sec.get("items", []):
            ws.cell(row=row, column=1, value=item["label"]).font = font(C["dark"])
            formula_cell(ws, row, 2, item.get("formula", ""))
            row += 1
        row += 1


# ════════════════════════════════════════════════════════════════════════
#  2. IMAGE BUILDER
# ════════════════════════════════════════════════════════════════════════

SIZE = 2000
MARGIN = 80

def _gradient_bg(color1: str, color2: str) -> Image.Image:
    img = Image.new("RGB", (SIZE, SIZE))
    c1 = hex2rgb(color1)
    c2 = hex2rgb(color2)
    draw = ImageDraw.Draw(img)
    for y in range(SIZE):
        t = y / SIZE
        r = int(c1[0] + (c2[0] - c1[0]) * t)
        g = int(c1[1] + (c2[1] - c1[1]) * t)
        b = int(c1[2] + (c2[2] - c1[2]) * t)
        draw.line([(0, y), (SIZE, y)], fill=(r, g, b))
    return img


def _rounded_rect(draw, x1, y1, x2, y2, radius=30, fill_color=None, outline=None, width=3):
    if fill_color:
        draw.rounded_rectangle([x1, y1, x2, y2], radius=radius, fill=fill_color, outline=outline, width=width)
    else:
        draw.rounded_rectangle([x1, y1, x2, y2], radius=radius, outline=outline, width=width)


def build_images(cfg: dict, slug: str) -> list:
    """Generate 5 Etsy product images. Returns list of file paths."""
    paths = []
    d = out_dir(slug)

    name = cfg.get("name", "Product")
    subtitle = cfg.get("subtitle", "Google Sheets Template")
    price = cfg.get("price", "$19")
    tabs_list = [t["name"] for t in cfg.get("tabs", [])]
    features = cfg.get("features", [])
    perfect_for = cfg.get("perfect_for", [])
    accent = cfg.get("accent_color", "4361EE")
    dark = cfg.get("dark_color", "0D1B2A")
    green = cfg.get("green_color", "06D6A0")
    yellow = cfg.get("yellow_color", "FFD166")
    red = cfg.get("red_color", "EF233C")

    # Image 1 — Hero
    img = _gradient_bg(dark, "1a2a4a")
    draw = ImageDraw.Draw(img)
    f_big   = try_font(96, bold=True)
    f_med   = try_font(56, bold=True)
    f_small = try_font(38)
    f_tiny  = try_font(30)

    draw.text((SIZE//2, 160), name.upper(), font=f_big, fill=hex2rgb(yellow), anchor="mm")
    draw.text((SIZE//2, 280), subtitle, font=f_med, fill=(255, 255, 255), anchor="mm")

    # KPI cards row
    kpis = cfg.get("kpis_preview", ["Revenue", "Clients", "Profit", "Tax"])
    card_w = 380
    total_w = len(kpis) * card_w + (len(kpis) - 1) * 20
    start_x = (SIZE - total_w) // 2
    for i, kpi in enumerate(kpis[:4]):
        cx = start_x + i * (card_w + 20)
        _rounded_rect(draw, cx, 380, cx + card_w, 560, 20,
                      fill_color=hex2rgb(accent) + (200,))
        draw.text((cx + card_w//2, 470), kpi, font=f_small,
                  fill=(255, 255, 255), anchor="mm")

    # Spreadsheet mockup lines
    mock_top = 620
    mock_left = 200
    mock_right = 1800
    draw.rectangle([mock_left, mock_top, mock_right, mock_top + 50],
                   fill=hex2rgb(accent))
    for i, col in enumerate(["#", "Name", "Status", "Value", "Date"]):
        x = mock_left + 20 + i * 320
        draw.text((x, mock_top + 25), col, font=f_tiny,
                  fill=(255, 255, 255), anchor="lm")
    for r in range(5):
        ry = mock_top + 50 + r * 44
        row_fill = (255, 255, 255, 15) if r % 2 == 0 else (255, 255, 255, 5)
        draw.rectangle([mock_left, ry, mock_right, ry + 44],
                       fill=hex2rgb("1e3050"))
        draw.line([(mock_left, ry + 44), (mock_right, ry + 44)],
                  fill=hex2rgb("2a4060"), width=1)

    # Price badge
    _rounded_rect(draw, SIZE - 380, SIZE - 260, SIZE - 80, SIZE - 80,
                  30, fill_color=hex2rgb(green))
    draw.text((SIZE - 230, SIZE - 195), "INSTANT", font=try_font(30, True),
              fill=hex2rgb(dark), anchor="mm")
    draw.text((SIZE - 230, SIZE - 150), "DOWNLOAD", font=try_font(30, True),
              fill=hex2rgb(dark), anchor="mm")
    draw.text((SIZE - 230, SIZE - 105), price, font=try_font(50, True),
              fill=hex2rgb(dark), anchor="mm")

    draw.text((SIZE//2, SIZE - 80), "NasriTools | nasritools.etsy.com",
              font=f_tiny, fill=hex2rgb("aaaaaa"), anchor="mm")

    p = str(d / f"{slug}_01_hero.jpg")
    img.save(p, "JPEG", quality=95)
    paths.append(p)

    # Image 2 — Tabs / What's Included
    img2 = _gradient_bg("0a0f1e", "111827")
    draw2 = ImageDraw.Draw(img2)
    draw2.text((SIZE//2, 100), "WHAT'S INSIDE", font=try_font(80, True),
               fill=hex2rgb(yellow), anchor="mm")
    draw2.text((SIZE//2, 180), f"{len(tabs_list)} Professional Tabs — Everything You Need",
               font=try_font(42), fill=(255, 255, 255), anchor="mm")

    tab_colors = [accent, green, red, yellow, "9B59B6", "E67E22", "1ABC9C", "E74C3C"]
    cols_count = 2
    card_w2 = 860
    card_h2 = 160
    gap = 30
    start_y = 260
    for i, tab_name in enumerate(tabs_list[:8]):
        col = i % cols_count
        row = i // cols_count
        cx = MARGIN + col * (card_w2 + gap)
        cy = start_y + row * (card_h2 + gap)
        tc = hex2rgb(tab_colors[i % len(tab_colors)])
        _rounded_rect(draw2, cx, cy, cx + card_w2, cy + card_h2,
                      20, fill_color=(tc[0], tc[1], tc[2]))
        draw2.text((cx + 30, cy + card_h2//2), tab_name,
                   font=try_font(44, True), fill=(255, 255, 255), anchor="lm")

    draw2.text((SIZE//2, SIZE - 80), "NasriTools | nasritools.etsy.com",
               font=try_font(30), fill=hex2rgb("aaaaaa"), anchor="mm")
    p2 = str(d / f"{slug}_02_included.jpg")
    img2.save(p2, "JPEG", quality=95)
    paths.append(p2)

    # Image 3 — How to Use (4 steps)
    img3 = _gradient_bg("0D1B2A", "162032")
    draw3 = ImageDraw.Draw(img3)
    draw3.text((SIZE//2, 100), "HOW TO USE", font=try_font(80, True),
               fill=hex2rgb(yellow), anchor="mm")
    draw3.text((SIZE//2, 190), "Ready in 3 minutes", font=try_font(42),
               fill=(200, 200, 200), anchor="mm")

    steps = cfg.get("how_to_use", [
        "Purchase & Download the file instantly",
        "Open in Google Sheets or Excel",
        "Fill in the YELLOW cells with your data",
        "Everything updates automatically!",
    ])
    step_colors = [accent, green, yellow, red]
    for i, step in enumerate(steps[:4]):
        y_center = 340 + i * 380
        circle_x = 250
        _rounded_rect(draw3, circle_x - 80, y_center - 80,
                      circle_x + 80, y_center + 80,
                      80, fill_color=hex2rgb(step_colors[i]))
        draw3.text((circle_x, y_center), str(i + 1),
                   font=try_font(72, True), fill=(255, 255, 255), anchor="mm")
        draw3.text((400, y_center), step, font=try_font(52, True),
                   fill=(255, 255, 255), anchor="lm")
        if i < 3:
            draw3.line([(circle_x, y_center + 80), (circle_x, y_center + 300)],
                       fill=hex2rgb("444444"), width=3)

    draw3.text((SIZE//2, SIZE - 80), "NasriTools | nasritools.etsy.com",
               font=try_font(30), fill=hex2rgb("aaaaaa"), anchor="mm")
    p3 = str(d / f"{slug}_03_how.jpg")
    img3.save(p3, "JPEG", quality=95)
    paths.append(p3)

    # Image 4 — Features grid
    img4 = _gradient_bg("111111", "1a1a2e")
    draw4 = ImageDraw.Draw(img4)
    draw4.text((SIZE//2, 100), "KEY FEATURES", font=try_font(80, True),
               fill=hex2rgb(green), anchor="mm")

    card_fw = 870
    card_fh = 200
    fgap = 30
    fstart_y = 200
    for i, feat in enumerate(features[:8]):
        col = i % 2
        row = i // 2
        fx = MARGIN + col * (card_fw + fgap)
        fy = fstart_y + row * (card_fh + fgap)
        _rounded_rect(draw4, fx, fy, fx + card_fw, fy + card_fh,
                      18, fill_color=hex2rgb("1e2a3a"))
        _rounded_rect(draw4, fx, fy, fx + 12, fy + card_fh,
                      0, fill_color=hex2rgb(accent))
        draw4.text((fx + 30, fy + card_fh//2 - 20), "✓",
                   font=try_font(48, True), fill=hex2rgb(green), anchor="lm")
        draw4.text((fx + 90, fy + card_fh//2), feat,
                   font=try_font(38), fill=(255, 255, 255), anchor="lm")

    draw4.text((SIZE//2, SIZE - 80), "NasriTools | nasritools.etsy.com",
               font=try_font(30), fill=hex2rgb("aaaaaa"), anchor="mm")
    p4 = str(d / f"{slug}_04_features.jpg")
    img4.save(p4, "JPEG", quality=95)
    paths.append(p4)

    # Image 5 — CTA / Perfect For
    img5 = _gradient_bg("0D1B2A", "0a0f1e")
    draw5 = ImageDraw.Draw(img5)
    draw5.text((SIZE//2, 110), "PERFECT FOR", font=try_font(80, True),
               fill=hex2rgb(yellow), anchor="mm")

    for i, item in enumerate(perfect_for[:6]):
        y5 = 250 + i * 220
        _rounded_rect(draw5, MARGIN, y5, SIZE - MARGIN, y5 + 180,
                      20, fill_color=hex2rgb("1a2a3a"))
        draw5.text((MARGIN + 40, y5 + 90), "→",
                   font=try_font(52, True), fill=hex2rgb(accent), anchor="lm")
        draw5.text((MARGIN + 120, y5 + 90), item,
                   font=try_font(46), fill=(255, 255, 255), anchor="lm")

    # Bottom CTA
    _rounded_rect(draw5, MARGIN, SIZE - 280, SIZE - MARGIN, SIZE - 100,
                  30, fill_color=hex2rgb(green))
    draw5.text((SIZE//2, SIZE - 190), f"INSTANT DOWNLOAD  ·  {price}  ·  YOURS FOREVER",
               font=try_font(52, True), fill=hex2rgb(dark), anchor="mm")

    p5 = str(d / f"{slug}_05_cta.jpg")
    img5.save(p5, "JPEG", quality=95)
    paths.append(p5)

    return paths


# ════════════════════════════════════════════════════════════════════════
#  3. PDF GUIDE BUILDER
# ════════════════════════════════════════════════════════════════════════

def build_pdf(cfg: dict, slug: str) -> str:
    styles = getSampleStyleSheet()
    ACCENT  = colors.HexColor(f"#{cfg.get('accent_color','4361EE')}")
    DARK    = colors.HexColor(f"#{cfg.get('dark_color','0D1B2A')}")
    GREEN   = colors.HexColor(f"#{cfg.get('green_color','06D6A0')}")
    YELLOW  = colors.HexColor(f"#{cfg.get('yellow_color','FFD166')}")

    def h1(t): return Paragraph(t, ParagraphStyle("h1", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=22, textColor=DARK,
        spaceAfter=6, spaceBefore=14, alignment=TA_CENTER))
    def h2(t): return Paragraph(t, ParagraphStyle("h2", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=14, textColor=ACCENT,
        spaceAfter=4, spaceBefore=10))
    def body(t): return Paragraph(t, ParagraphStyle("body", parent=styles["Normal"],
        fontName="Helvetica", fontSize=11, textColor=DARK,
        spaceAfter=4, leading=16))
    def bul(t): return Paragraph(f"• {t}", ParagraphStyle("bul", parent=styles["Normal"],
        fontName="Helvetica", fontSize=11, textColor=DARK,
        leftIndent=16, spaceAfter=3, leading=15))

    path = str(out_dir(slug) / f"{slug}_guide.pdf")
    doc = SimpleDocTemplate(path, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=20*mm, bottomMargin=20*mm)
    story = []

    # Cover
    story.append(Spacer(1, 10*mm))
    story.append(Paragraph(
        f'<font color="#{cfg.get("accent_color","4361EE")}">{cfg["name"].upper()}</font>',
        ParagraphStyle("cov", parent=styles["Normal"],
            fontName="Helvetica-Bold", fontSize=26, alignment=TA_CENTER,
            textColor=ACCENT, spaceAfter=6)))
    story.append(Paragraph(cfg.get("subtitle", ""), ParagraphStyle("sub",
        parent=styles["Normal"], fontName="Helvetica", fontSize=14,
        alignment=TA_CENTER, textColor=colors.HexColor("#888888"), spaceAfter=4)))
    story.append(Paragraph("NasriTools | nasritools.etsy.com",
        ParagraphStyle("store", parent=styles["Normal"], fontName="Helvetica-Bold",
        fontSize=11, alignment=TA_CENTER, textColor=DARK, spaceAfter=10)))
    story.append(HRFlowable(width="100%", thickness=2, color=ACCENT))
    story.append(Spacer(1, 8*mm))

    # Quick Start
    story.append(h2("Quick Start (3 Steps)"))
    steps_data = cfg.get("quick_start", [
        ("Download", "Download the file after purchase."),
        ("Open", "Open in Google Sheets (File → Import) or Excel."),
        ("Fill & Go", "Fill YELLOW cells — everything updates automatically."),
    ])
    tdata = [[Paragraph(f"<b>Step {i+1}</b>", styles["Normal"]),
              Paragraph(f"<b>{t}</b><br/>{d}", styles["Normal"])]
             for i, (t, d) in enumerate(steps_data)]
    tbl = Table(tdata, colWidths=[30*mm, 130*mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#F8F9FA")),
        ("BACKGROUND", (0,0), (0,-1), ACCENT),
        ("TEXTCOLOR",  (0,0), (0,-1), colors.white),
        ("FONTNAME",   (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE",   (0,0), (-1,-1), 11),
        ("ALIGN",      (0,0), (0,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#DDDDDD")),
        ("TOPPADDING", (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 6*mm))

    # Tabs overview
    tabs = cfg.get("tabs", [])
    if tabs:
        story.append(h2(f"What's Inside ({len(tabs)} Tabs)"))
        tab_data = [[Paragraph(f"<b>{t['name']}</b>", styles["Normal"]),
                     Paragraph(t.get("description", ""), styles["Normal"])]
                    for t in tabs]
        tbl2 = Table(tab_data, colWidths=[52*mm, 108*mm])
        tbl2.setStyle(TableStyle([
            ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
            ("FONTSIZE", (0,0), (-1,-1), 10),
            ("VALIGN",   (0,0), (-1,-1), "TOP"),
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, colors.HexColor("#F8F9FA")]),
            ("GRID",     (0,0), (-1,-1), 0.5, colors.HexColor("#DDDDDD")),
            ("TOPPADDING", (0,0), (-1,-1), 7),
            ("BOTTOMPADDING", (0,0), (-1,-1), 7),
            ("LEFTPADDING", (0,0), (-1,-1), 8),
        ]))
        story.append(tbl2)
        story.append(Spacer(1, 6*mm))

    # Features
    features = cfg.get("features", [])
    if features:
        story.append(h2("Key Features"))
        for f in features:
            story.append(bul(f))
        story.append(Spacer(1, 6*mm))

    # Color guide
    story.append(h2("Color Guide"))
    story.append(bul("<b>Yellow cells</b> = Enter your data here"))
    story.append(bul("<b>Blue/colored cells</b> = Formulas (do not edit)"))
    story.append(bul("<b>Green rows</b> = Completed / Paid / Active"))
    story.append(bul("<b>Red rows</b> = Overdue / Unpaid / Urgent"))
    story.append(Spacer(1, 6*mm))

    # Tips
    tips = cfg.get("tips", [])
    if tips:
        story.append(h2("Pro Tips"))
        for tip in tips:
            story.append(bul(tip))

    story.append(Spacer(1, 8*mm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#DDDDDD")))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph(
        "Thank you for purchasing from NasriTools!<br/>"
        "Questions? Contact us at nasritools.etsy.com",
        ParagraphStyle("footer", parent=styles["Normal"],
            fontName="Helvetica", fontSize=10, alignment=TA_CENTER,
            textColor=colors.HexColor("#888888"))))

    doc.build(story)
    return path


# ════════════════════════════════════════════════════════════════════════
#  4. ETSY LISTING TEXT BUILDER
# ════════════════════════════════════════════════════════════════════════

def build_listing(cfg: dict, slug: str) -> str:
    name       = cfg["name"]
    subtitle   = cfg.get("subtitle", "")
    price      = cfg.get("price", "$19")
    price_note = cfg.get("price_note", "")
    tabs       = cfg.get("tabs", [])
    features   = cfg.get("features", [])
    perfect    = cfg.get("perfect_for", [])
    tags       = cfg.get("tags", [])
    title      = cfg.get("listing_title", f"{name} Google Sheets Template | {subtitle}")

    desc_header = cfg.get("listing_emoji", "🔵")

    lines = []
    lines.append("━" * 50)
    lines.append(f"ETSY LISTING — {name.upper()}")
    lines.append("━" * 50)
    lines.append(f"\nTITLE (140 chars max):\n{title[:140]}")
    lines.append(f"\nPRICE: {price}")
    lines.append(f"\nCATEGORY:\n{cfg.get('category', 'Craft Supplies & Tools → Templates → Spreadsheets')}")
    lines.append("\n─── DESCRIPTION " + "─" * 34)
    lines.append(f"\n{desc_header} {name.upper()}\n{subtitle}\n")
    lines.append(cfg.get("description_intro", ""))
    lines.append("\n━━━━━━━━━━━━━━━━━━━")
    lines.append("✅ WHAT'S INCLUDED")
    lines.append("━━━━━━━━━━━━━━━━━━━\n")
    lines.append("📥 Instant Download:")
    lines.append(f"• {slug}.xlsx (works in Google Sheets & Excel)")
    lines.append(f"• PDF Setup Guide\n")
    if tabs:
        lines.append(f"{len(tabs)} Professional Tabs:")
        for t in tabs:
            lines.append(f"{t.get('emoji','📋')} {t['name']} — {t.get('description','')}")
    lines.append("\n━━━━━━━━━━━━━━━━━━━")
    lines.append("🎯 PERFECT FOR")
    lines.append("━━━━━━━━━━━━━━━━━━━")
    for p in perfect:
        lines.append(f"• {p}")
    lines.append("\n━━━━━━━━━━━━━━━━━━━")
    lines.append("⚡ KEY FEATURES")
    lines.append("━━━━━━━━━━━━━━━━━━━")
    for f in features:
        lines.append(f"✔ {f}")
    lines.append("\n━━━━━━━━━━━━━━━━━━━")
    lines.append("🚀 HOW TO USE")
    lines.append("━━━━━━━━━━━━━━━━━━━")
    how = cfg.get("how_to_use", [
        "Purchase & Download instantly",
        "Open in Google Sheets or Excel",
        "Fill in the YELLOW cells",
        "Watch everything update automatically!",
    ])
    for i, step in enumerate(how, 1):
        lines.append(f"{i}️⃣ {step}")
    lines.append("\n━━━━━━━━━━━━━━━━━━━")
    lines.append("📦 INSTANT DOWNLOAD")
    lines.append("━━━━━━━━━━━━━━━━━━━")
    lines.append("After purchase you'll receive:")
    lines.append("• .xlsx file (Google Sheets & Microsoft Excel)")
    lines.append("• PDF User Guide")
    lines.append("\n⭐ Questions? Message us — we respond within 24 hours.")
    lines.append("\n─── TAGS (13 tags) " + "─" * 31)
    for tag in tags[:13]:
        lines.append(tag)
    if price_note:
        lines.append(f"\n─── NOTES\n{price_note}")
    lines.append("━" * 50)

    text = "\n".join(lines)
    path = str(out_dir(slug) / f"{slug}_listing.txt")
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(text)
    return path


# ════════════════════════════════════════════════════════════════════════
#  5. MAIN FACTORY FUNCTION
# ════════════════════════════════════════════════════════════════════════

def build_product(cfg: dict) -> dict:
    """
    Master function. Accepts a product config dict, returns paths dict.
    cfg must have at minimum: 'name', 'slug'
    """
    slug = cfg["slug"]
    print(f"\n{'═'*50}")
    print(f"  Building: {cfg['name']}")
    print(f"{'═'*50}")

    print("  [1/4] Excel template...")
    xlsx = build_excel(cfg, slug)
    print(f"        ✅ {xlsx}")

    print("  [2/4] Product images...")
    imgs = build_images(cfg, slug)
    print(f"        ✅ {len(imgs)} images")

    print("  [3/4] PDF guide...")
    pdf = build_pdf(cfg, slug)
    print(f"        ✅ {pdf}")

    print("  [4/4] Etsy listing text...")
    listing = build_listing(cfg, slug)
    print(f"        ✅ {listing}")

    result = {
        "slug": slug,
        "name": cfg["name"],
        "excel": xlsx,
        "images": imgs,
        "pdf": pdf,
        "listing": listing,
        "built_at": datetime.now().isoformat(),
    }

    summary_path = str(out_dir(slug) / "build_result.json")
    with open(summary_path, "w") as fh:
        json.dump(result, fh, indent=2)

    print(f"\n  Done! All files in: output/{slug}/")
    return result


if __name__ == "__main__":
    # Quick self-test with a sample product
    sample = {
        "slug": "budget_tracker",
        "name": "Personal Budget Tracker",
        "subtitle": "Monthly & Annual Google Sheets Template",
        "price": "$12",
        "accent_color": "4361EE",
        "dark_color": "0D1B2A",
        "green_color": "06D6A0",
        "yellow_color": "FFD166",
        "red_color": "EF233C",
        "listing_title": "Personal Budget Tracker Google Sheets | Monthly Budget Planner | Finance Spreadsheet",
        "listing_emoji": "💰",
        "category": "Craft Supplies & Tools → Templates → Spreadsheets",
        "description_intro": "Take full control of your money with this powerful yet simple budget tracker.",
        "kpis_preview": ["Income", "Expenses", "Savings", "Balance"],
        "tabs": [
            {"name": "Dashboard",    "emoji": "🏠", "color": "0D1B2A", "type": "dashboard",
             "description": "Monthly overview: income, expenses, savings rate, balance.",
             "kpis": [
                 {"label": "Total Income",   "formula": "=SUM(Income!B:B)"},
                 {"label": "Total Expenses", "formula": "=SUM(Expenses!C:C)"},
                 {"label": "Net Savings",    "formula": "=Dashboard!B4-Dashboard!C4"},
                 {"label": "Savings Rate",   "formula": '=TEXT(Dashboard!D4/Dashboard!B4,"0%")'},
             ]},
            {"name": "Income",       "emoji": "💵", "color": "06D6A0", "type": "table",
             "description": "Log all income sources with date, category, and amount.",
             "columns": [
                 {"name": "Date",     "width": 14, "sample": "2024-01-01"},
                 {"name": "Source",   "width": 20, "sample": "Salary"},
                 {"name": "Category", "width": 18, "dropdown": ["Salary","Freelance","Business","Investment","Other"], "sample": "Salary"},
                 {"name": "Amount",   "width": 14, "sample": "3000"},
                 {"name": "Notes",    "width": 26, "sample": "Monthly salary"},
             ], "sample_rows": 8},
            {"name": "Expenses",     "emoji": "💳", "color": "EF233C", "type": "table",
             "description": "Track all expenses by category with budget vs actual.",
             "columns": [
                 {"name": "Date",     "width": 14, "sample": "2024-01-05"},
                 {"name": "Item",     "width": 22, "sample": "Rent"},
                 {"name": "Category", "width": 18, "dropdown": ["Housing","Food","Transport","Health","Entertainment","Other"], "sample": "Housing"},
                 {"name": "Amount",   "width": 14, "sample": "800"},
                 {"name": "Budget",   "width": 14, "sample": "850"},
                 {"name": "Notes",    "width": 22, "sample": "Monthly rent"},
             ], "sample_rows": 10},
            {"name": "Savings Goals","emoji": "🎯", "color": "4361EE", "type": "table",
             "description": "Set savings goals and track progress automatically.",
             "columns": [
                 {"name": "Goal",     "width": 24, "sample": "Emergency Fund"},
                 {"name": "Target",   "width": 14, "sample": "5000"},
                 {"name": "Saved",    "width": 14, "sample": "1200"},
                 {"name": "Progress", "width": 14, "formula": '=TEXT(D{ROW}/C{ROW},"0%")'},
                 {"name": "Deadline", "width": 16, "sample": "2024-12-31"},
             ], "sample_rows": 5},
            {"name": "Monthly View", "emoji": "📅", "color": "9B59B6", "type": "table",
             "description": "Month-by-month comparison of income, expenses, and savings.",
             "columns": [
                 {"name": "Month",    "width": 16, "sample": "January"},
                 {"name": "Income",   "width": 14, "sample": "3000"},
                 {"name": "Expenses", "width": 14, "sample": "2100"},
                 {"name": "Savings",  "width": 14, "formula": "=C{ROW}-D{ROW}"},
                 {"name": "Rate",     "width": 12, "formula": '=TEXT(E{ROW}/C{ROW},"0%")'},
             ], "sample_rows": 12},
            {"name": "Reports",      "emoji": "📊", "color": "E67E22", "type": "reports",
             "description": "Auto summary: top expense categories, savings trend, annual totals.",
             "sections": [
                 {"title": "Annual Summary", "items": [
                     {"label": "Total Income",   "formula": "=SUM(Income!B:B)"},
                     {"label": "Total Expenses", "formula": "=SUM(Expenses!D:D)"},
                     {"label": "Total Savings",  "formula": "=SUM(Income!B:B)-SUM(Expenses!D:D)"},
                 ]},
             ]},
        ],
        "features": [
            "Auto savings rate calculation",
            "Budget vs Actual comparison (turns red if over budget)",
            "12-month view with trend tracking",
            "Savings goals with progress bars",
            "Income categorization (Salary, Freelance, Business, etc.)",
            "Expense categorization (Housing, Food, Transport, etc.)",
            "Annual totals auto-calculated",
            "No subscriptions — yours forever",
        ],
        "perfect_for": [
            "Anyone wanting to control their monthly spending",
            "People saving for a specific goal (house, car, trip)",
            "Freelancers tracking variable income",
            "Families managing household budgets",
            "Students learning to budget for the first time",
            "Anyone tired of expensive budgeting apps",
        ],
        "how_to_use": [
            "Purchase & download the .xlsx file instantly",
            "Open in Google Sheets or Excel",
            "Fill YELLOW cells with your income and expenses",
            "Dashboard shows your financial picture automatically",
        ],
        "tips": [
            "Update expenses daily for best accuracy.",
            "Set a savings goal first — it motivates consistent tracking.",
            "Change the tax rate in the Settings cell if you need it.",
            "Duplicate the file each year for annual archives.",
            "Color code categories with Google Sheets conditional formatting.",
        ],
        "tags": [
            "budget tracker google sheets",
            "monthly budget template",
            "personal finance spreadsheet",
            "expense tracker excel",
            "savings goal tracker",
            "budget planner template",
            "household budget sheet",
            "income expense tracker",
            "financial planner google",
            "money management template",
            "monthly expenses sheet",
            "budget spreadsheet instant",
            "personal budget tracker",
        ],
    }

    build_product(sample)
