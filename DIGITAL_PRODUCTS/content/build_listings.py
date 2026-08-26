import openpyxl, json, os

PLAN = r'C:\Users\nasri\Desktop\PREMIUM_TEMPLATES\🚀_خطة_الإطلاق.xlsx'
CONT = r'C:\Users\nasri\Desktop\excel\محتوى_المنتجات_جاهز.xlsx'
ETSY = r'C:\Users\nasri\AppData\Local\Temp\claude\C--Users-nasri--claude\b6fd1f80-a8e7-48d6-9bfb-ff88677c10b1\scratchpad\etsy'

# descriptions keyed by old product name
wb2 = openpyxl.load_workbook(CONT, data_only=True)
descs = {}
for r in wb2['Descriptions'].iter_rows(min_row=3, values_only=True):
    if r[0]: descs[r[0].strip()] = r[1]

# map new product -> old description key + assets
MAP = {
 'Trading Journal':        ('Options Trading Journal',            'PREMIUM_Trading_Journal.xlsx',  '01_Trading_Journal.png','included_trading.png','demo_trading.mp4'),
 'Business KPI Dashboard': ('Business KPI Dashboard',             'PREMIUM_Business_KPI.xlsx',     '02_Business_KPI.png','included_business.png','demo_business.mp4'),
 'Invoice & Client Tracker':('Freelancer Invoice Tracker',        'PREMIUM_Invoice_Client.xlsx',   '03_Invoice_Client.png','included_invoice.png','demo_invoice.mp4'),
 'Monthly Budget Tracker': ('Monthly Budget & Expense Tracker',   'PREMIUM_Budget_Tracker.xlsx',   '04_Budget_Tracker.png','included_budget.png','demo_budget.mp4'),
 'Habit Tracker':          ('Habit Tracker & 30-Day Challenge',   'PREMIUM_Habit_Tracker.xlsx',    '05_Habit_Tracker.png','included_habit.png','demo_habit.mp4'),
 'Student Planner + GPA':  ('Student Planner & GPA Calculator',   'PREMIUM_Student_Planner.xlsx',  '06_Student_Planner.png','included_student.png','demo_student.mp4'),
 'Meal Planner + Grocery': ('Meal Planner & Grocery List',        'PREMIUM_Meal_Planner.xlsx',     '07_Meal_Planner.png','included_meal.png','demo_meal.mp4'),
}

wb = openpyxl.load_workbook(PLAN, data_only=True)
out = []
for r in wb['Launch Plan'].iter_rows(min_row=3, values_only=True):
    if not r[0] or 'BUNDLE' in str(r[1]): continue
    name = str(r[1]).strip()
    if name not in MAP: 
        print('SKIP:', name); continue
    dkey, xlsx, img1, img2, vid = MAP[name]
    imgs = [os.path.join(ETSY,i) for i in (img1, img2, '00_How_It_Works.png') if os.path.exists(os.path.join(ETSY,i))]
    out.append({
        'product': name,
        'title': str(r[5]).strip(),
        'tags': [t.strip() for t in str(r[6]).split(',')][:13],
        'price': float(r[3]),
        'description': descs.get(dkey,'').strip(),
        'file': os.path.join(ETSY, xlsx),
        'images': imgs,
        'video': os.path.join(ETSY, vid) if os.path.exists(os.path.join(ETSY,vid)) else None,
    })

with open('listings.json','w',encoding='utf-8') as f:
    json.dump(out,f,ensure_ascii=False,indent=1)

print(f'BUILT {len(out)} listings\n')
for L in out:
    print(f"[{L['product']}] ${L['price']}")
    print(f"  title  : {len(L['title'])} chars {'OK' if len(L['title'])<=140 else 'TOO LONG!'}")
    print(f"  tags   : {len(L['tags'])} {'OK' if all(len(t)<=20 for t in L['tags']) else 'SOME >20 CHARS!'}")
    print(f"  desc   : {len(L['description'])} chars")
    print(f"  images : {len(L['images'])} | video: {'yes' if L['video'] else 'no'} | file: {os.path.basename(L['file'])} {'OK' if os.path.exists(L['file']) else 'MISSING'}")
