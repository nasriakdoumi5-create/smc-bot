"""
build_starter_v2.py
Builds a REAL Starter Budget Tracker — a template worth paying for:
  • START HERE tab with styled instructions
  • Dashboard: this-month KPIs + 12-month table + category breakdown
  • Income / Expenses tabs with dropdowns & realistic sample data
  • Budget tab: budget vs actual with red over-budget alerts
  • Savings Goals with progress bars
All formulas are Google Sheets compatible.
Run:  python build_starter_v2.py
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.chart import BarChart, Reference
from pathlib import Path
from datetime import date

def d(s):
    y, m, dd = s.split("-")
    return date(int(y), int(m), int(dd))

OUT = Path(__file__).parent / "output" / "starter_budget_tracker" / "starter_budget_tracker.xlsx"

# ── palette ────────────────────────────────────────────────────────────
DARK    = "1A1A2E"
ORANGE  = "F97316"
GREEN   = "22C55E"
RED     = "EF4444"
CREAM   = "FFF7F0"
LIGHT   = "F4F4F6"
WHITE   = "FFFFFF"

F_TITLE  = Font(name="Arial", bold=True, size=20, color=WHITE)
F_SUB    = Font(name="Arial", size=11, color="DDDDDD")
F_H      = Font(name="Arial", bold=True, size=12, color=WHITE)
F_KPI_L  = Font(name="Arial", bold=True, size=11, color=WHITE)
F_KPI_V  = Font(name="Arial", bold=True, size=18, color=WHITE)
F_TH     = Font(name="Arial", bold=True, size=10, color=WHITE)
F_TD     = Font(name="Arial", size=10, color="222222")
F_NOTE   = Font(name="Arial", italic=True, size=10, color="777777")
F_STEP   = Font(name="Arial", bold=True, size=12, color=ORANGE)
F_BODY   = Font(name="Arial", size=11, color="333333")

FILL_DARK   = PatternFill("solid", fgColor=DARK)
FILL_ORANGE = PatternFill("solid", fgColor=ORANGE)
FILL_GREEN  = PatternFill("solid", fgColor=GREEN)
FILL_CREAM  = PatternFill("solid", fgColor=CREAM)
FILL_LIGHT  = PatternFill("solid", fgColor=LIGHT)

THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")

CATEGORIES = ["Housing", "Food", "Transport", "Health", "Fun", "Shopping", "Bills", "Other"]
SOURCES    = ["Salary", "Freelance", "Side Hustle", "Gift", "Other"]
EUR = '"€"#,##0.00'

def th_row(ws, row, headers, widths=None, fill=FILL_DARK):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F_TH; c.fill = fill; c.alignment = CENTER; c.border = BORDER
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── 1. START HERE ──────────────────────────────────────────────────────
ws = wb.active
ws.title = "START HERE"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for col, w in zip("BCDEF", [16, 30, 30, 30, 16]):
    ws.column_dimensions[col].width = w

ws.merge_cells("B2:F3")
c = ws["B2"]; c.value = "💰  STARTER BUDGET TRACKER"; c.font = F_TITLE
c.fill = FILL_DARK; c.alignment = CENTER
ws.merge_cells("B4:F4")
c = ws["B4"]; c.value = "Welcome! You're 3 steps away from knowing exactly where your money goes."
c.font = F_SUB; c.fill = FILL_DARK; c.alignment = CENTER

steps = [
    ("STEP 1 — Add your income", 'Open the "Income" tab and log everything you earn: salary, freelance, side money. Just type the date, amount and pick a source.'),
    ("STEP 2 — Log your expenses", 'Open the "Expenses" tab every time you spend. Pick a category from the dropdown — this powers all the charts.'),
    ("STEP 3 — Watch the Dashboard", 'The "Dashboard" tab updates itself: this month\'s income, spending, what\'s left, and your savings rate. Set limits in "Budget" and goals in "Savings Goals".'),
]
r = 6
for title, body in steps:
    ws.merge_cells(f"B{r}:F{r}")
    c = ws[f"B{r}"]; c.value = title; c.font = F_STEP
    ws.merge_cells(f"B{r+1}:F{r+1}")
    c = ws[f"B{r+1}"]; c.value = body; c.font = F_BODY
    ws.row_dimensions[r+1].height = 30
    r += 3

tips = [
    "TIP: Delete the sample rows once you've seen how it works — select them and press Delete.",
    "TIP: On Google Sheets? File → Make a copy first, then edit your own copy.",
    "TIP: Everything recalculates automatically. You never need to touch a formula.",
    "Need help? Message NasriTools on Etsy — I reply fast. Enjoy! 🧡",
]
for tip in tips:
    ws.merge_cells(f"B{r}:F{r}")
    c = ws[f"B{r}"]; c.value = tip; c.font = F_NOTE
    r += 2

# ── 2. Income ──────────────────────────────────────────────────────────
ws = wb.create_sheet("Income")
ws.sheet_view.showGridLines = False
th_row(ws, 1, ["Date", "Amount", "Source", "Notes"], [14, 14, 18, 32])
income_rows = [
    ("2026-07-01", 2200, "Salary",      "Monthly salary"),
    ("2026-07-05",  350, "Freelance",   "Logo project"),
    ("2026-07-12",  120, "Side Hustle", "Etsy shop sales"),
    ("2026-07-19",   60, "Other",       "Sold old phone"),
    ("2026-07-26",  200, "Freelance",   "Consulting call"),
]
for i, (dt, a, s, n) in enumerate(income_rows, 2):
    c = ws.cell(row=i, column=1, value=d(dt)); c.font = F_TD; c.number_format = "yyyy-mm-dd"
    c = ws.cell(row=i, column=2, value=a); c.font = F_TD; c.number_format = EUR
    ws.cell(row=i, column=3, value=s).font = F_TD
    ws.cell(row=i, column=4, value=n).font = F_TD
    for col in range(1, 5):
        ws.cell(row=i, column=col).border = BORDER
        if i % 2 == 0:
            ws.cell(row=i, column=col).fill = FILL_CREAM

dv = DataValidation(type="list", formula1='"' + ",".join(SOURCES) + '"', allow_blank=True)
ws.add_data_validation(dv); dv.add("C2:C500")

# ── 3. Expenses ────────────────────────────────────────────────────────
ws = wb.create_sheet("Expenses")
ws.sheet_view.showGridLines = False
th_row(ws, 1, ["Date", "Item", "Category", "Amount", "Notes"], [14, 24, 16, 14, 30])
expense_rows = [
    ("2026-07-01", "Rent",             "Housing",   750, "Monthly rent"),
    ("2026-07-02", "Groceries",        "Food",       82, "Weekly shop"),
    ("2026-07-03", "Metro card",       "Transport",  40, "Monthly pass"),
    ("2026-07-05", "Pharmacy",         "Health",     18, ""),
    ("2026-07-07", "Cinema",           "Fun",        24, "Weekend movie"),
    ("2026-07-09", "Groceries",        "Food",       76, "Weekly shop"),
    ("2026-07-10", "Electricity bill", "Bills",      65, ""),
    ("2026-07-12", "New shoes",        "Shopping",   59, "Summer sale"),
    ("2026-07-14", "Restaurant",       "Food",       45, "Dinner out"),
    ("2026-07-16", "Groceries",        "Food",       80, "Weekly shop"),
    ("2026-07-18", "Internet bill",    "Bills",      35, ""),
    ("2026-07-21", "Fuel",             "Transport",  55, ""),
    ("2026-07-23", "Groceries",        "Food",       74, "Weekly shop"),
    ("2026-07-25", "Gym",              "Health",     30, "Monthly"),
    ("2026-07-28", "Gift",             "Other",      25, "Friend's birthday"),
]
for i, (dt, item, cat, a, n) in enumerate(expense_rows, 2):
    c = ws.cell(row=i, column=1, value=d(dt)); c.font = F_TD; c.number_format = "yyyy-mm-dd"
    ws.cell(row=i, column=2, value=item).font = F_TD
    ws.cell(row=i, column=3, value=cat).font = F_TD
    c = ws.cell(row=i, column=4, value=a); c.font = F_TD; c.number_format = EUR
    ws.cell(row=i, column=5, value=n).font = F_TD
    for col in range(1, 6):
        ws.cell(row=i, column=col).border = BORDER
        if i % 2 == 0:
            ws.cell(row=i, column=col).fill = FILL_CREAM

dv = DataValidation(type="list", formula1='"' + ",".join(CATEGORIES) + '"', allow_blank=True)
ws.add_data_validation(dv); dv.add("C2:C500")

# ── 4. Dashboard ───────────────────────────────────────────────────────
ws = wb.create_sheet("Dashboard", 1)
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for col, w in zip("BCDEFG", [16, 15, 15, 15, 15, 15]):
    ws.column_dimensions[col].width = w

ws.merge_cells("B2:G2")
c = ws["B2"]; c.value = "📊  MY BUDGET DASHBOARD"; c.font = Font(name="Arial", bold=True, size=16, color=WHITE)
c.fill = FILL_DARK; c.alignment = CENTER
ws.row_dimensions[2].height = 30

# KPI cards (this month, auto by TODAY)
kpis = [
    ("B", "THIS MONTH IN",
     '=SUMPRODUCT((TEXT(Income!$A$2:$A$500,"yyyy-mm")=TEXT(TODAY(),"yyyy-mm"))*Income!$B$2:$B$500)', FILL_GREEN),
    ("D", "THIS MONTH OUT",
     '=SUMPRODUCT((TEXT(Expenses!$A$2:$A$500,"yyyy-mm")=TEXT(TODAY(),"yyyy-mm"))*Expenses!$D$2:$D$500)', PatternFill("solid", fgColor=RED)),
    ("F", "LEFT TO SPEND", "=B5-D5", FILL_ORANGE),
]
for col, label, formula, fill in kpis:
    col2 = chr(ord(col) + 1)
    ws.merge_cells(f"{col}4:{col2}4"); ws.merge_cells(f"{col}5:{col2}5")
    c = ws[f"{col}4"]; c.value = label; c.font = F_KPI_L; c.fill = fill; c.alignment = CENTER
    c2 = ws[f"{col}5"]; c2.value = formula; c2.font = F_KPI_V; c2.fill = fill
    c2.alignment = CENTER; c2.number_format = EUR
    ws.row_dimensions[5].height = 28

# 12-month table
ws["B8"] = "MONTH-BY-MONTH (2026)"; ws["B8"].font = Font(name="Arial", bold=True, size=12, color=DARK)
th_row(ws, 9, [""] * 0, None)  # noop keeps helper import used
for i, h in enumerate(["Month", "Income", "Expenses", "Saved", "Rate"], 2):
    c = ws.cell(row=9, column=i, value=h)
    c.font = F_TH; c.fill = FILL_DARK; c.alignment = CENTER; c.border = BORDER

months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
for i, m in enumerate(months):
    r = 10 + i
    ym = f"2026-{i+1:02d}"
    ws.cell(row=r, column=2, value=m).font = F_TD
    ws.cell(row=r, column=3,
        value=f'=SUMPRODUCT((TEXT(Income!$A$2:$A$500,"yyyy-mm")="{ym}")*Income!$B$2:$B$500)')
    ws.cell(row=r, column=4,
        value=f'=SUMPRODUCT((TEXT(Expenses!$A$2:$A$500,"yyyy-mm")="{ym}")*Expenses!$D$2:$D$500)')
    ws.cell(row=r, column=5, value=f"=C{r}-D{r}")
    ws.cell(row=r, column=6, value=f'=IFERROR(TEXT(E{r}/C{r},"0%"),"—")')
    for col in range(2, 7):
        cell = ws.cell(row=r, column=col)
        cell.border = BORDER
        if col in (3, 4, 5): cell.number_format = EUR
        if cell.font is None or not cell.font.name: cell.font = F_TD
        if r % 2 == 0: cell.fill = FILL_CREAM

# category breakdown
ws["B24"] = "SPENDING BY CATEGORY (all time)"; ws["B24"].font = Font(name="Arial", bold=True, size=12, color=DARK)
for i, h in enumerate(["Category", "Spent"], 2):
    c = ws.cell(row=25, column=i, value=h)
    c.font = F_TH; c.fill = FILL_DARK; c.alignment = CENTER; c.border = BORDER
for i, cat in enumerate(CATEGORIES):
    r = 26 + i
    ws.cell(row=r, column=2, value=cat).font = F_TD
    c = ws.cell(row=r, column=3, value=f'=SUMIFS(Expenses!$D$2:$D$500,Expenses!$C$2:$C$500,B{r})')
    c.number_format = EUR; c.font = F_TD
    ws.cell(row=r, column=2).border = BORDER; c.border = BORDER
ws.conditional_formatting.add(
    f"C26:C{25+len(CATEGORIES)}",
    DataBarRule(start_type="num", start_value=0, end_type="max",
                color=ORANGE, showValue=True))

# chart: monthly income vs expenses
chart = BarChart()
chart.title = "Income vs Expenses"
chart.height, chart.width = 8, 16
data = Reference(ws, min_col=3, max_col=4, min_row=9, max_row=21)
cats = Reference(ws, min_col=2, min_row=10, max_row=21)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws.add_chart(chart, "H4")

# ── 5. Budget ──────────────────────────────────────────────────────────
ws = wb.create_sheet("Budget")
ws.sheet_view.showGridLines = False
th_row(ws, 1, ["Category", "Monthly Budget", "Spent This Month", "Left", "Status"],
       [16, 16, 18, 14, 14])
defaults = {"Housing": 800, "Food": 400, "Transport": 120, "Health": 80,
            "Fun": 100, "Shopping": 100, "Bills": 150, "Other": 60}
for i, cat in enumerate(CATEGORIES):
    r = 2 + i
    ws.cell(row=r, column=1, value=cat).font = F_TD
    b = ws.cell(row=r, column=2, value=defaults[cat]); b.number_format = EUR; b.font = F_TD
    s = ws.cell(row=r, column=3,
        value=f'=SUMPRODUCT((TEXT(Expenses!$A$2:$A$500,"yyyy-mm")=TEXT(TODAY(),"yyyy-mm"))*(Expenses!$C$2:$C$500=A{r})*Expenses!$D$2:$D$500)')
    s.number_format = EUR; s.font = F_TD
    l = ws.cell(row=r, column=4, value=f"=B{r}-C{r}"); l.number_format = EUR; l.font = F_TD
    st = ws.cell(row=r, column=5, value=f'=IF(C{r}>B{r},"⚠ OVER","✓ OK")')
    st.font = F_TD; st.alignment = CENTER
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = BORDER
        if r % 2 == 0: ws.cell(row=r, column=col).fill = FILL_CREAM

last = 1 + len(CATEGORIES)
ws.conditional_formatting.add(f"D2:D{last}",
    CellIsRule(operator="lessThan", formula=["0"],
               fill=PatternFill("solid", fgColor="FEE2E2"),
               font=Font(color=RED, bold=True)))
ws.conditional_formatting.add(f"E2:E{last}",
    CellIsRule(operator="equal", formula=['"⚠ OVER"'],
               fill=PatternFill("solid", fgColor="FEE2E2"),
               font=Font(color=RED, bold=True)))

# ── 6. Savings Goals ───────────────────────────────────────────────────
ws = wb.create_sheet("Savings Goals")
ws.sheet_view.showGridLines = False
th_row(ws, 1, ["Goal", "Target", "Saved So Far", "Progress", "Note"],
       [24, 14, 16, 14, 26])
goals = [
    ("Emergency Fund", 3000, 850,  "3 months of expenses"),
    ("Summer Trip",    1200, 400,  "July next year"),
    ("New Laptop",      900, 620,  ""),
]
for i, (g, t, s, n) in enumerate(goals, 2):
    ws.cell(row=i, column=1, value=g).font = F_TD
    c = ws.cell(row=i, column=2, value=t); c.number_format = EUR; c.font = F_TD
    c = ws.cell(row=i, column=3, value=s); c.number_format = EUR; c.font = F_TD
    p = ws.cell(row=i, column=4, value=f"=IFERROR(C{i}/B{i},0)")
    p.number_format = "0%"; p.font = F_TD; p.alignment = CENTER
    ws.cell(row=i, column=5, value=n).font = F_TD
    for col in range(1, 6):
        ws.cell(row=i, column=col).border = BORDER
ws.conditional_formatting.add("D2:D50",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                color=GREEN, showValue=True))

wb.save(OUT)
size = OUT.stat().st_size
print(f"Saved: {OUT}  ({size:,} bytes)")
print("Sheets:", wb.sheetnames)
